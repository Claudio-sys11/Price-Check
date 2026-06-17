"""
EcountERP 재고현황 + Wizfasta 가격비교 - Windows 데스크톱 앱 (Tkinter GUI).

탭 구성:
  ① 재고현황 조회 : EcountERP Open API 로 재고현황 조회 → 표 표시 / CSV 내보내기
  ② 가격비교      : Wizfasta 판매상품(JSON)과 EcountERP 재고를 품목코드로 결합 비교

설정은 %APPDATA%\\EcountInventory\\config.json 에 저장됩니다.
"""

from __future__ import annotations

import json
import os
import re
import sys
import threading
import time
import tkinter as tk
import tkinter.font as tkfont
from tkinter import ttk, messagebox, filedialog

from tksheet import Sheet

from ecount_api import EcountClient, EcountApiError
import compare as cmp
import updater
import backend
from version import APP_VERSION

APP_NAME = "EcountInventory"
DEFAULT_GITHUB_REPO = "Claudio-sys11/Price-Check"   # 자동 업데이트 기본 저장소

# 인증 정보 고정값
FIXED_COM_CODE = "188894"          # 회사코드(고정·수정 불가)
FIXED_USER_ID = "THEFEELKOREA"     # 사용자ID(고정·수정 불가)
DEFAULT_API_CERT_KEY = "28ac7027d054c443cb50b538fc5063f058"   # API 인증키(기본값·수정 가능)


def resource_path(rel: str) -> str:
    """PyInstaller 빌드/개발 환경 모두에서 번들 리소스 경로를 해결한다."""
    base = getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base, rel)


# ===== 고해상도(DPI) 스케일링 =====
# UI_SCALE = 화면 DPI / 96  (96dpi=100%→1.0, 144dpi=150%→1.5). 1920×1080·2560×1600 등
# 서로 다른 배율에서도 동일한 비율·선명함으로 보이도록 커스텀 위젯/창 크기를 곱해 키운다.
UI_SCALE = 1.0


def enable_dpi_awareness():
    """프로세스를 DPI 인식으로 만들어 고DPI에서 흐릿하지 않게 한다(Tk 생성 전 호출)."""
    try:
        import ctypes
        try:
            ctypes.windll.shcore.SetProcessDpiAwareness(1)   # System DPI aware
        except Exception:  # noqa: BLE001
            ctypes.windll.user32.SetProcessDPIAware()
    except Exception:  # noqa: BLE001
        pass


def init_ui_scale(root):
    """루트 창의 실제 DPI로 UI_SCALE 을 계산한다."""
    global UI_SCALE
    try:
        dpi = float(root.winfo_fpixels("1i"))
        UI_SCALE = max(1.0, min(3.0, dpi / 96.0))
    except Exception:  # noqa: BLE001
        UI_SCALE = 1.0
    return UI_SCALE


def sc(v):
    """픽셀 값을 현재 UI 배율로 환산(정수)."""
    return int(round(v * UI_SCALE))


def _scale_photo(img):
    """tk.PhotoImage 를 UI_SCALE 에 맞춰 정수 zoom/subsample 로 확대(고DPI 대응)."""
    if img is None or abs(UI_SCALE - 1.0) < 0.01:
        return img
    try:
        from fractions import Fraction
        fr = Fraction(UI_SCALE).limit_denominator(8)
        return img.zoom(fr.numerator).subsample(fr.denominator)
    except Exception:  # noqa: BLE001
        return img

# ===== 디자인 팔레트 =====
FONT = "맑은 고딕"   # Malgun Gothic — 한글에 최적화된 깔끔한 글자체
BG = "#f3f4f6"          # 전체 배경(연한 회색)
CARD = "#ffffff"        # 카드/표 배경
TEXT = "#1f2937"        # 본문 텍스트
SUBTLE = "#374151"      # 라벨프레임 제목
MUTED = "#6b7280"       # 보조 텍스트
BORDER = "#d1d5db"      # 테두리
ACCENT = "#14b8a6"      # 강조(민트/틸)
ACCENT_ACTIVE = "#0d9488"
ACCENT_SOFT = "#cbf5ec"  # 민트 연한톤(헤더 보조 텍스트 등)
HEADING_BG = "#e0f7f1"  # 표 헤더 배경(연한 민트)
HEADING_FG = "#0f766e"  # 표 헤더 글자(진한 틸)
ROW_ALT = "#f5fbf9"     # 표 줄무늬(홀수행, 민트 기운)
SUBTOTAL_BG = "#d7f5ed"  # 소계 행 강조(민트)
SELECT_BG = "#f0e0a8"   # 표 선택행(샴페인 골드 — 민트 테마와 뚜렷이 구분되는 음영)
SELECT_FG = "#0f1f1c"   # 선택행 글자

# 프리미엄(고급) 톤
HEADER_DARK = "#0b5c54"   # 헤더 그라데이션 진한 쪽(딥 틸)
GOLD = "#c9a227"          # 고급 포인트(골드)
GOLD_SOFT = "#ecd9a0"
INK = "#0f1f1c"           # 깊은 먹빛 텍스트
HAIRLINE = "#e7edeb"      # 얇은 구분선

# 라운드(동글) UI 색상
TAB_INACTIVE = "#dbe7e3"        # 비활성 탭 알약 배경(연한 민트회색)
TAB_INACTIVE_HOVER = "#cbe0d9"  # 비활성 탭 hover
BTN_GRAY = "#e5e7eb"            # 보조 버튼
BTN_GRAY_ACTIVE = "#d1d5db"
BTN_GRAY_DISABLED = "#f0f1f3"
BTN_GRAY_FG = TEXT
BTN_GRAY_FG_DISABLED = "#9ca3af"
DANGER = "#f87171"             # 중단 버튼(연한 빨강)
DANGER_ACTIVE = "#ef4444"
DANGER_DISABLED = "#f6cfcf"
DIFF_BG = "#fff1f0"            # 가격비교: 단가차이 행
DIFF_FG = "#b91c1c"
UNMATCH_BG = "#eef1f4"         # 가격비교: 미매칭 행
UNMATCH_FG = "#6b7280"

MONEY_COLS = {"입고단가", "총단가",
              "파스타원가", "평균원가(ERP)", "차이",
              "파스타재고", "실재고(ERP)", "재고차이"}  # 우측정렬(금액·수량) 컬럼

SPINNER = "⠋⠙⠹⠸⠼⠴⠦⠧⠇⠏"   # 실시간 활동 스피너

# Wizfasta 원가 가져오기 진행 체크리스트 단계
WIZ_STEPS = [
    ("start", "Chrome 시작"),
    ("login", "Wizfasta 로그인"),
    ("query", "상품DB 조회 (일반상품·재고≥1)"),
    ("download", "상품 데이터 수집"),
    ("parse", "데이터 파싱"),
    ("match", "모델명 매칭"),
]


def app_data_dir() -> str:
    base = os.environ.get("APPDATA") or os.path.expanduser("~")
    path = os.path.join(base, APP_NAME)
    os.makedirs(path, exist_ok=True)
    return path


CONFIG_PATH = os.path.join(app_data_dir(), "config.json")
PRICE_CACHE_PATH = os.path.join(app_data_dir(), "price_cache.json")


def load_price_cache() -> dict:
    try:
        with open(PRICE_CACHE_PATH, encoding="utf-8") as f:
            return json.load(f) or {}
    except (OSError, json.JSONDecodeError):
        return {}


def save_price_cache(cache: dict) -> None:
    try:
        with open(PRICE_CACHE_PATH, "w", encoding="utf-8") as f:
            json.dump(cache, f, ensure_ascii=False)
    except OSError:
        pass


DAILY_STATUS_PATH = os.path.join(app_data_dir(), "daily_status.json")


def load_daily_status() -> list:
    """일일현황 이력(원가비교 대시보드 집계)을 일자 목록으로 반환."""
    try:
        with open(DAILY_STATUS_PATH, encoding="utf-8") as f:
            data = json.load(f)
            return data if isinstance(data, list) else []
    except (OSError, json.JSONDecodeError):
        return []


def save_daily_status(records: list) -> None:
    try:
        with open(DAILY_STATUS_PATH, "w", encoding="utf-8") as f:
            json.dump(records, f, ensure_ascii=False, indent=2)
    except OSError:
        pass


def _sort_key(value) -> tuple:
    """정렬 키: 숫자로 해석되면 숫자 기준, 아니면 문자 기준(콤마 제거)."""
    s = str(value).replace(",", "").strip()
    try:
        return (0, float(s))
    except ValueError:
        return (1, str(value))


# 의류 사이즈 순서(S < M < L ...). 숫자 사이즈는 작은 것부터.
_SIZE_ORDER = {
    "XXS": 0, "XS": 1, "S": 2, "M": 3, "L": 4, "XL": 5,
    "XXL": 6, "2XL": 6, "XXXL": 7, "3XL": 7, "4XL": 8,
    "U": 50, "F": 51, "FREE": 51, "ONE": 51, "ONESIZE": 51,
}


def _size_key(value) -> tuple:
    """사이즈 정렬 키: 글자 사이즈는 S,M,L 순 / 숫자 사이즈는 작은→큰 / 기타는 텍스트."""
    s = str(value).strip().upper()
    if s in _SIZE_ORDER:
        return (0, _SIZE_ORDER[s], 0.0, "")
    if re.fullmatch(r"[0-9]+(\.[0-9]+)?", s):     # 순수 숫자 사이즈
        return (1, 0, float(s), "")
    return (2, 0, 0.0, s)                          # 그 외 텍스트


def fill_tree(tree: ttk.Treeview, rows: list[dict],
              sort_callback=None, sort_col: str | None = None,
              sort_desc: bool = False) -> None:
    import tkinter.font as tkfont

    tree.delete(*tree.get_children())
    if not rows:
        tree["columns"] = ()
        return
    headers: list[str] = []
    for r in rows:
        for k in r.keys():
            if not k.startswith("_") and k not in headers:
                headers.append(k)
    tree["columns"] = headers

    # 내용에 맞춘 열 너비 자동 최적화
    fnt = tkfont.Font(family=FONT, size=10)
    fnt_b = tkfont.Font(family=FONT, size=10, weight="bold")
    for h in headers:
        arrow = (" ▼" if sort_desc else " ▲") if h == sort_col else ""
        head_text = f"{h}{arrow}"
        if sort_callback is not None:
            tree.heading(h, text=head_text, anchor="center",
                         command=lambda c=h: sort_callback(c))
        else:
            tree.heading(h, text=head_text, anchor="center")
        # 헤더(화살표 여유 포함) + 모든 셀 중 최대 픽셀폭
        w = fnt_b.measure(head_text + "  ")
        for r in rows:
            w = max(w, fnt.measure(str(r.get(h, ""))))
        w = min(max(w + 26, 60), 420)   # 여백 + 최소/최대 제한
        # 금액 컬럼은 우측정렬, 그 외는 가운데
        anchor = "e" if h in MONEY_COLS else "center"
        tree.column(h, width=w, anchor=anchor, stretch=False)

    tree.tag_configure("subtotal", background=SUBTOTAL_BG, font=(FONT, 10, "bold"))
    tree.tag_configure("grandtotal", background="#9fe3d6", foreground="#0f3d36",
                       font=(FONT, 10, "bold"))
    tree.tag_configure("odd", background=ROW_ALT)
    tree.tag_configure("even", background="white")
    # 가격비교 행 색상(정렬 그룹 구분): 단가차이 / 미매칭 / 동일
    tree.tag_configure("cmp_diff", background=DIFF_BG, foreground=DIFF_FG)
    tree.tag_configure("cmp_unmatched", background=UNMATCH_BG, foreground=UNMATCH_FG)
    tree.tag_configure("cmp_same", background="white")
    i = 0
    for r in rows:
        if r.get("_grand"):
            tags: tuple = ("grandtotal",)
        elif r.get("_subtotal"):
            tags = ("subtotal",)
        elif r.get("_tag") == "diff":
            tags = ("cmp_diff",)
        elif r.get("_tag") == "unmatched":
            tags = ("cmp_unmatched",)
        elif r.get("_tag") == "same":
            tags = ("cmp_same",)
        else:
            tags = ("odd",) if (i % 2) else ("even",)
            i += 1
        tree.insert("", "end", values=[r.get(h, "") for h in headers], tags=tags)


# ===================== 동글동글(라운드) UI 컴포넌트 =====================
def _hex_rgb(h):
    h = h.lstrip("#")
    return tuple(int(h[i:i + 2], 16) for i in (0, 2, 4))


def _paint_h_gradient(canvas, w, h, c1, c2, step=2):
    """캔버스에 가로 방향 그라데이션을 그린다(c1 → c2)."""
    r1, g1, b1 = _hex_rgb(c1)
    r2, g2, b2 = _hex_rgb(c2)
    w = max(1, int(w))
    for x in range(0, w, step):
        t = x / (w - 1) if w > 1 else 0
        r = int(r1 + (r2 - r1) * t)
        g = int(g1 + (g2 - g1) * t)
        b = int(b1 + (b2 - b1) * t)
        canvas.create_line(x, 0, x, h, fill=f"#{r:02x}{g:02x}{b:02x}", width=step)


def _round_rect_points(x1, y1, x2, y2, r):
    """둥근 사각형을 그릴 다각형 좌표(smooth=True 와 함께 사용)."""
    r = max(0, min(r, (x2 - x1) / 2, (y2 - y1) / 2))
    return [
        x1 + r, y1,  x2 - r, y1,  x2, y1,
        x2, y1 + r,  x2, y2 - r,  x2, y2,
        x2 - r, y2,  x1 + r, y2,  x1, y2,
        x1, y2 - r,  x1, y1 + r,  x1, y1,
    ]


class RoundedButton(tk.Canvas):
    """동글동글한 라운드 버튼(Canvas 기반) — ttk.Button 대체.

    지원: command, state('normal'/'disabled'), text 변경, hover 효과.
    부모 배경색(bg)을 받아 모서리가 자연스럽게 비치도록 한다.
    """

    def __init__(self, parent, text="", command=None, *, bg=BG,
                 fill=ACCENT, fill_active=ACCENT_ACTIVE, fill_disabled="#a7ddd4",
                 fg="white", fg_disabled="#e6fffa",
                 font=(FONT, 10, "bold"), height=38, radius=19, padx=20, minwidth=0):
        height, radius, padx, minwidth = sc(height), sc(radius), sc(padx), sc(minwidth)
        super().__init__(parent, bg=bg, highlightthickness=0, bd=0, height=height)
        self._command = command
        self._fill, self._fill_active, self._fill_disabled = fill, fill_active, fill_disabled
        self._fg, self._fg_disabled = fg, fg_disabled
        self._font = font
        self._radius, self._h = radius, height
        self._state = "normal"
        self._text = text
        self._hover = False
        weight = font[2] if len(font) > 2 else "normal"
        self._fnt = tkfont.Font(family=font[0], size=font[1], weight=weight)
        w = max(minwidth, self._fnt.measure(text) + 2 * padx)
        super().configure(width=w)
        self.bind("<Configure>", lambda e: self._draw())
        self.bind("<Button-1>", self._on_click)
        self.bind("<Enter>", self._on_enter)
        self.bind("<Leave>", self._on_leave)
        self._draw()

    def _cur_w(self):
        w = self.winfo_width()
        return w if w > 1 else int(self["width"])

    def _draw(self):
        self.delete("all")
        w, h = self._cur_w(), self._h
        if self._state == "disabled":
            fillc, fgc = self._fill_disabled, self._fg_disabled
        elif self._hover:
            fillc, fgc = self._fill_active, self._fg
        else:
            fillc, fgc = self._fill, self._fg
        self.create_polygon(_round_rect_points(1, 1, w - 1, h - 1, self._radius),
                            fill=fillc, outline="", smooth=True)
        self.create_text(w // 2, h // 2, text=self._text, fill=fgc, font=self._font)

    def _on_click(self, _e):
        if self._state != "disabled" and self._command:
            self._command()

    def _on_enter(self, _e):
        if self._state != "disabled":
            self._hover = True
            self.configure(cursor="hand2")
            self._draw()

    def _on_leave(self, _e):
        self._hover = False
        self.configure(cursor="")
        self._draw()

    def configure(self, cnf=None, **kw):   # ttk.Button 처럼 state/text/command 처리
        if cnf and isinstance(cnf, dict):
            kw.update(cnf)
        redraw = False
        if "state" in kw:
            self._state = kw.pop("state")
            self._hover = False
            redraw = True
        if "text" in kw:
            self._text = kw.pop("text")
            redraw = True
        if "command" in kw:
            self._command = kw.pop("command")
        if kw:
            super().configure(**kw)
        if redraw:
            self._draw()

    config = configure


class RoundedTabBar(tk.Frame):
    """알약형(pill) 탭 바. 탭별 크기·위치(좌/우)를 지정할 수 있다.

    tabs: [(key, label, opts), ...]
        opts: {"side": "left"|"right", "height", "font", "padx", "radius"}
    활성 여부는 색상으로 구분한다. on_select(key) 콜백으로 선택을 알린다.
    """

    def __init__(self, parent, tabs, on_select, *, bg=BG,
                 active_fill=ACCENT, active_fg="white",
                 inactive_fill=TAB_INACTIVE, inactive_fg=MUTED,
                 hover_fill=TAB_INACTIVE_HOVER):
        super().__init__(parent, bg=bg)
        self._on_select = on_select
        self._active_fill, self._active_fg = active_fill, active_fg
        self._inactive_fill, self._inactive_fg = inactive_fill, inactive_fg
        self._hover_fill = hover_fill
        self._pills = {}
        self._active = None
        left_count = 0
        for key, label, opts in tabs:
            side = opts.get("side", "left")
            h = sc(opts.get("height", 44))
            font = opts.get("font", (FONT, 11, "bold"))
            padx = sc(opts.get("padx", 28))
            radius = sc(opts.get("radius", 22))
            fnt = tkfont.Font(family=font[0], size=font[1],
                              weight=font[2] if len(font) > 2 else "normal")
            w = fnt.measure(label) + 2 * padx
            c = tk.Canvas(self, bg=bg, highlightthickness=0, bd=0, width=w, height=h)
            gap = sc(8)
            if side == "right":
                c.pack(side="right", padx=(gap, 0))
            else:
                c.pack(side="left", padx=(0 if left_count == 0 else gap, 0))
                left_count += 1
            c.bind("<Button-1>", lambda e, k=key: self._select(k))
            c.bind("<Enter>", lambda e, k=key: self._hover(k, True))
            c.bind("<Leave>", lambda e, k=key: self._hover(k, False))
            self._pills[key] = {"c": c, "label": label, "w": w, "h": h,
                                "radius": radius, "font": font}
        if tabs:
            self.select(tabs[0][0])

    def _paint(self, key, fill, fg):
        p = self._pills[key]
        c = p["c"]
        c.delete("all")
        c.create_polygon(_round_rect_points(1, 1, p["w"] - 1, p["h"] - 1, p["radius"]),
                         fill=fill, outline="", smooth=True)
        c.create_text(p["w"] // 2, p["h"] // 2, text=p["label"], fill=fg, font=p["font"])

    def _render(self, key, hover=False):
        if key == self._active:
            self._paint(key, self._active_fill, self._active_fg)
        else:
            self._paint(key, self._hover_fill if hover else self._inactive_fill,
                        self._inactive_fg)

    def _hover(self, key, on):
        self._pills[key]["c"].configure(cursor="hand2" if on else "")
        if key != self._active:
            self._render(key, hover=on)

    def _select(self, key):
        self.select(key)
        if self._on_select:
            self._on_select(key)

    def select(self, key):
        self._active = key
        for k in self._pills:
            self._render(k)


def accent_button(parent, text, command, *, bg=BG, **kw):
    """민트 강조 라운드 버튼."""
    return RoundedButton(parent, text, command, bg=bg, fill=ACCENT,
                         fill_active=ACCENT_ACTIVE, fill_disabled="#a7ddd4",
                         fg="white", fg_disabled="#e6fffa", **kw)


def gray_button(parent, text, command, *, bg=BG, **kw):
    """보조(회색) 라운드 버튼."""
    return RoundedButton(parent, text, command, bg=bg, fill=BTN_GRAY,
                         fill_active=BTN_GRAY_ACTIVE, fill_disabled=BTN_GRAY_DISABLED,
                         fg=BTN_GRAY_FG, fg_disabled=BTN_GRAY_FG_DISABLED, **kw)


class StatChip(tk.Canvas):
    """둥근 통계 칩 — 큰 숫자 + 라벨. command 가 있으면 클릭 가능(필터)하며,
    선택 시 골드 테두리로 활성 표시한다."""

    def __init__(self, parent, label, *, bg=BG, fill="#e0f7f1", fg="#0f766e",
                 cw=148, ch=60, radius=20, command=None):
        cw, ch, radius = sc(cw), sc(ch), sc(radius)
        super().__init__(parent, bg=bg, highlightthickness=0, bd=0, width=cw, height=ch)
        self._label = label
        self._fill, self._fg = fill, fg
        self._radius = radius
        self._cw, self._ch = cw, ch
        self._value = "—"
        self._pct = ""
        self._command = command
        self._active = False
        self.bind("<Configure>", lambda e: self._draw())
        if command:
            self.bind("<Button-1>", lambda e: command())
            self.bind("<Enter>", lambda e: self.configure(cursor="hand2"))
            self.bind("<Leave>", lambda e: self.configure(cursor=""))
        self._draw()

    def set_value(self, value):
        self._value = str(value)
        self._draw()

    def set_pct(self, pct_text):
        self._pct = str(pct_text or "")
        self._draw()

    def set_active(self, active):
        self._active = bool(active)
        self._draw()

    def _draw(self):
        self.delete("all")
        self.create_polygon(_round_rect_points(1, 1, self._cw - 1, self._ch - 1, self._radius),
                            fill=self._fill, outline="", smooth=True)
        if self._active:   # 선택 표시(골드 테두리)
            self.create_polygon(_round_rect_points(2, 2, self._cw - 2, self._ch - 2, self._radius),
                                fill="", outline=GOLD, width=max(2, sc(3)), smooth=True)
        self.create_text(self._cw // 2, self._ch // 2 - sc(9), text=self._value,
                         fill=self._fg, font=(FONT, 18, "bold"))
        label = self._label + (f"  {self._pct}" if self._pct else "")
        self.create_text(self._cw // 2, self._ch // 2 + sc(15), text=label,
                         fill=self._fg, font=(FONT, 9, "bold"))


EXPORT_FONT_SIZE = 10   # 내보낸 엑셀의 글자 크기(pt)


def export_rows_excel(rows: list[dict], initial: str = "export.xlsx",
                      font_size: int = EXPORT_FONT_SIZE) -> None:
    """조회된 항목을 엑셀(.xlsx)로 저장한다. 모든 글자에 size=font_size(기본 10) 적용.

    확장자를 .csv 로 지정하면 CSV(서식 없음)로 저장한다.
    """
    if not rows:
        pmsg.showwarning("데이터 없음", "내보낼 데이터가 없습니다.")
        return
    path = filedialog.asksaveasfilename(
        title="엑셀로 저장", defaultextension=".xlsx",
        filetypes=[("Excel 파일", "*.xlsx"), ("CSV 파일", "*.csv"), ("모든 파일", "*.*")],
        initialfile=initial,
    )
    if not path:
        return

    headers: list[str] = []
    for r in rows:
        for k in r.keys():
            if not k.startswith("_") and k not in headers:
                headers.append(k)

    try:
        if path.lower().endswith(".csv"):
            import csv
            with open(path, "w", newline="", encoding="utf-8-sig") as f:
                w = csv.DictWriter(f, fieldnames=headers, extrasaction="ignore")
                w.writeheader()
                w.writerows(rows)
        else:
            import openpyxl
            from openpyxl.styles import Font, Alignment
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "데이터"
            ws.append(headers)
            for r in rows:
                ws.append([r.get(h, "") for h in headers])

            # 글자 크기 10 적용(헤더는 굵게), 가운데 정렬
            base = Font(name="맑은 고딕", size=font_size)
            head = Font(name="맑은 고딕", size=font_size, bold=True)
            center = Alignment(horizontal="center", vertical="center")
            for ci in range(1, len(headers) + 1):
                ws.cell(row=1, column=ci).font = head
                ws.cell(row=1, column=ci).alignment = center
            for ri in range(2, len(rows) + 2):
                for ci in range(1, len(headers) + 1):
                    ws.cell(row=ri, column=ci).font = base

            # 열 너비 대략 맞춤
            for ci, h in enumerate(headers, start=1):
                maxlen = max([len(str(h))] + [len(str(r.get(h, ""))) for r in rows])
                ws.column_dimensions[get_column_letter(ci)].width = min(max(maxlen + 2, 8), 50)

            ws.freeze_panes = "A2"   # 머리글 고정
            wb.save(path)
    except Exception as exc:  # noqa: BLE001
        pmsg.showerror("저장 실패", str(exc))
        return
    pmsg.showinfo("저장 완료", f"{len(rows)}건을 저장했습니다.\n{path}")


# 하위 호환(기존 호출명 유지)
export_rows_csv = export_rows_excel


class Splash(tk.Toplevel):
    """실행 시 표시되는 프리미엄 로딩 화면(둥근 모서리).

    EcountERP·Wizfasta 로고를 배치하고, 업데이트 설치 시 진행률(0~100%)을 표시.
    Windows 의 -transparentcolor 로 카드 밖(모서리)을 투명 처리해 둥근 창을 만든다.
    """

    KEY = "#FF00FE"   # 투명 처리용 키 컬러(콘텐츠에 쓰지 않는 색)

    def __init__(self, parent, status: str = "로딩 중…"):
        super().__init__(parent)
        self.overrideredirect(True)
        self._cw, self._ch = sc(560), sc(340)
        sw, sh = self.winfo_screenwidth(), self.winfo_screenheight()
        self.geometry(f"{self._cw}x{self._ch}+{(sw - self._cw) // 2}+{(sh - self._ch) // 2}")
        try:
            self.attributes("-topmost", True)
        except tk.TclError:
            pass
        bg = "white"
        try:
            self.attributes("-transparentcolor", self.KEY)   # 모서리 투명 → 둥근 창
            self.configure(bg=self.KEY)
            bg = self.KEY
        except tk.TclError:
            self.configure(bg="white")

        w, h = self._cw, self._ch
        c = tk.Canvas(self, width=w, height=h, bg=bg, highlightthickness=0, bd=0)
        c.pack(fill="both", expand=True)
        self.canvas = c

        # 로고 이미지
        self._img_ec = None
        self._img_wz = None
        try:
            self._img_ec = _scale_photo(tk.PhotoImage(file=resource_path("assets/ecount_logo_splash.png")))
            self._img_wz = _scale_photo(tk.PhotoImage(file=resource_path("assets/wizfasta_logo_splash.png")))
        except tk.TclError:
            self._img_ec = self._img_wz = None

        # 카드(둥근 흰색) + 얇은 보더 + 상단 골드 악센트
        m, rad = sc(8), sc(32)
        c.create_polygon(_round_rect_points(m, m, w - m, h - m, rad),
                         fill="white", outline="", smooth=True)
        c.create_polygon(_round_rect_points(w // 2 - sc(28), m + sc(16), w // 2 + sc(28), m + sc(21), sc(2)),
                         fill=GOLD, outline="", smooth=True)

        # 로고 행
        cy = sc(124)
        if self._img_ec is not None and self._img_wz is not None:
            ew, wzw = self._img_ec.width(), self._img_wz.width()
            gap, xw = sc(24), sc(18)
            total = ew + gap + xw + gap + wzw
            x = (w - total) // 2
            c.create_image(x + ew // 2, cy, image=self._img_ec)
            c.create_text(x + ew + gap + xw // 2, cy, text="×", fill="#c2cecb", font=(FONT, 16))
            c.create_image(x + ew + gap + xw + gap + wzw // 2, cy, image=self._img_wz)
        else:
            c.create_text(w // 2, cy, text="EcountERP  ×  Wizfasta",
                          fill=ACCENT_ACTIVE, font=(FONT, 18, "bold"))

        # 구분선 + 버전 · 게시자
        c.create_line(w // 2 - sc(160), sc(166), w // 2 + sc(160), sc(166), fill=HAIRLINE)
        c.create_text(w // 2, sc(188), text=f"VERSION {APP_VERSION}", fill=GOLD,
                      font=(FONT, 9, "bold"))
        c.create_text(w // 2, sc(207), text="T H E   F E E L   K O R E A   C O . , L T D .",
                      fill="#aab4b1", font=(FONT, 8))
        c.create_text(w // 2, sc(223), text="Created by Claudio Lim",
                      fill="#c2c9c6", font=(FONT, 8))

        # 상태 + 스피너(단일 텍스트, 가운데)
        self._status_text = status
        self._status_id = c.create_text(w // 2, sc(254), text=f"{SPINNER[0]}   {status}",
                                        fill="#5b6b67", font=(FONT, 10))

        # 진행률 바(다운로드/설치 시 생성)
        self._pct = 0.0
        self._bar_made = False
        self._pct_id = None
        self._bw, self._bh = sc(340), sc(10)
        self._bx = (w - self._bw) // 2
        self._by = sc(284)

        self._anim_on = True
        self._i = 0
        self._tick()
        try:
            self.update_idletasks()
            self.lift()
        except tk.TclError:
            pass

    def set_status(self, msg: str) -> None:
        self._status_text = msg
        try:
            self.canvas.itemconfigure(
                self._status_id, text=f"{SPINNER[self._i % len(SPINNER)]}   {msg}")
        except tk.TclError:
            pass

    def set_progress(self, pct: float) -> None:
        self._pct = max(0.0, min(100.0, float(pct)))
        c = self.canvas
        try:
            if not self._bar_made:
                c.create_polygon(
                    _round_rect_points(self._bx, self._by, self._bx + self._bw,
                                       self._by + self._bh, self._bh // 2),
                    fill="#edf1f0", outline="", smooth=True, tags=("bartrack",))
                self._pct_id = c.create_text(
                    self._bx + self._bw + sc(28), self._by + self._bh // 2,
                    text="0%", fill=ACCENT_ACTIVE, font=(FONT, 9, "bold"))
                self._bar_made = True
            c.delete("barfill")
            fw = int(self._bw * self._pct / 100)
            if fw >= 2:
                rr = min(self._bh / 2, fw / 2)
                c.create_polygon(
                    _round_rect_points(self._bx, self._by, self._bx + fw,
                                       self._by + self._bh, rr),
                    fill=ACCENT, outline="", smooth=True, tags=("barfill",))
            c.itemconfigure(self._pct_id, text=f"{int(round(self._pct))}%")
        except tk.TclError:
            pass

    def _tick(self) -> None:
        if not self._anim_on:
            return
        self._i += 1
        try:
            self.canvas.itemconfigure(
                self._status_id, text=f"{SPINNER[self._i % len(SPINNER)]}   {self._status_text}")
        except tk.TclError:
            pass
        self.after(120, self._tick)

    def close(self) -> None:
        self._anim_on = False
        try:
            self.destroy()
        except tk.TclError:
            pass


class PremiumDialog(tk.Toplevel):
    """둥근 모서리 프리미엄 모달 다이얼로그(스플래시와 동일 디자인).

    표준 messagebox 를 대체하며 info/warning/error/ask(예·아니오) 종류를 지원한다.
    """

    KEY = "#FF00FE"

    def __init__(self, parent, title, message, kind="info", ask=False):
        super().__init__(parent)
        self.overrideredirect(True)
        self.result = False
        pal = {
            "info": (ACCENT, ACCENT_ACTIVE),
            "warning": (GOLD, "#a9851d"),
            "error": ("#ef4444", "#dc2626"),
            "ask": (ACCENT, ACCENT_ACTIVE),
        }
        accent, accent_dk = pal.get(kind, pal["info"])
        w = sc(440)
        wrap = w - sc(64)
        bg = "white"
        try:
            self.attributes("-topmost", True)
            self.attributes("-transparentcolor", self.KEY)
            self.configure(bg=self.KEY)
            bg = self.KEY
        except tk.TclError:
            self.configure(bg="white")

        c = tk.Canvas(self, width=w, height=sc(600), bg=bg, highlightthickness=0, bd=0)
        c.pack(fill="both", expand=True)

        # 메시지 텍스트(먼저 그려 높이 측정)
        mid = c.create_text(sc(32), sc(84), anchor="nw", text=str(message), fill="#374151",
                            font=(FONT, 10), width=wrap)
        bb = c.bbox(mid)
        h = max(sc(150), (bb[3] if bb else sc(110)) + sc(74))

        try:
            parent.update_idletasks()
            px, py = parent.winfo_rootx(), parent.winfo_rooty()
            pw, ph = parent.winfo_width(), parent.winfo_height()
            x, y = px + (pw - w) // 2, py + (ph - h) // 2
        except Exception:  # noqa: BLE001
            sw, sh = self.winfo_screenwidth(), self.winfo_screenheight()
            x, y = (sw - w) // 2, (sh - h) // 2
        self.geometry(f"{w}x{h}+{max(0, x)}+{max(0, y)}")
        c.config(height=h)

        # 카드(텍스트 뒤로) + 상단 악센트 + 제목 + 구분선
        m, rad = sc(6), sc(26)
        card = c.create_polygon(_round_rect_points(m, m, w - m, h - m, rad),
                                fill="white", outline="", smooth=True)
        acc = c.create_polygon(_round_rect_points(m, m, w - m, m + sc(6), sc(3)),
                               fill=accent, outline="", smooth=True)
        c.tag_lower(card, mid)
        c.tag_lower(acc, mid)
        c.create_text(sc(32), sc(44), anchor="w", text=str(title), fill=INK, font=(FONT, 13, "bold"))
        c.create_oval(w - sc(46), sc(38), w - sc(34), sc(50), fill=accent, outline="")
        c.create_line(sc(28), sc(64), w - sc(28), sc(64), fill=HAIRLINE)

        # 버튼
        bar = tk.Frame(self, bg="white")
        bar.place(x=w - sc(28), y=h - sc(30), anchor="e")

        def mk(text, cmd, fill, fdk, fg="white"):
            return RoundedButton(bar, text, cmd, bg="white", fill=fill, fill_active=fdk,
                                 fill_disabled=fill, fg=fg, fg_disabled=fg,
                                 height=34, radius=17, padx=18)

        if ask:
            mk("예", self._yes, accent, accent_dk).pack(side="right")
            mk("아니오", self._no, BTN_GRAY, BTN_GRAY_ACTIVE, fg=TEXT).pack(side="right", padx=(0, 10))
            self.bind("<Return>", lambda e: self._yes())
            self.bind("<Escape>", lambda e: self._no())
        else:
            mk("확인", self._ok, accent, accent_dk).pack(side="right")
            self.bind("<Return>", lambda e: self._ok())
            self.bind("<Escape>", lambda e: self._ok())

        try:
            self.transient(parent)
        except tk.TclError:
            pass
        try:
            self.update_idletasks()
            self.lift()
            self.focus_force()
            self.grab_set()
        except tk.TclError:
            pass
        self.wait_window(self)

    def _ok(self):
        self.result = True
        self._close()

    def _yes(self):
        self.result = True
        self._close()

    def _no(self):
        self.result = False
        self._close()

    def _close(self):
        try:
            self.grab_release()
        except tk.TclError:
            pass
        try:
            self.destroy()
        except tk.TclError:
            pass


def _active_root():
    getter = getattr(tk, "_get_default_root", None)
    if getter:
        try:
            return getter()
        except Exception:  # noqa: BLE001
            pass
    return getattr(tk, "_default_root", None)


class _PMsg:
    """messagebox 호환 shim — 둥근 프리미엄 다이얼로그로 표시."""

    @staticmethod
    def showinfo(title, message, **kw):
        PremiumDialog(_active_root(), title, message, "info")
        return "ok"

    @staticmethod
    def showwarning(title, message, **kw):
        PremiumDialog(_active_root(), title, message, "warning")
        return "ok"

    @staticmethod
    def showerror(title, message, **kw):
        PremiumDialog(_active_root(), title, message, "error")
        return "ok"

    @staticmethod
    def askyesno(title, message, **kw):
        return PremiumDialog(_active_root(), title, message, "ask", ask=True).result


pmsg = _PMsg()


class PremiumMenu(tk.Toplevel):
    """둥근 모서리 프리미엄 컨텍스트(우클릭) 메뉴.

    items: [(label, command), ...] — label 이 "-" 이면 구분선.
    호버 강조, 클릭 시 실행, 바깥 클릭/ESC 로 닫힘.
    """

    KEY = "#FF00FE"

    def __init__(self, parent, items, x, y):
        super().__init__(parent)
        self.overrideredirect(True)
        self._items = items
        self._row_h, self._sep_h, self._pad = sc(32), sc(9), sc(6)
        self._hover = -1

        fnt = tkfont.Font(family=FONT, size=10)
        labels = [t for t, _ in items if t != "-"]
        self._mw = (max((fnt.measure(t) for t in labels), default=sc(120))) + sc(56)
        h = self._pad * 2
        for t, _cmd in items:
            h += self._sep_h if t == "-" else self._row_h
        self._mh = h

        sw, sh = self.winfo_screenwidth(), self.winfo_screenheight()
        x = min(int(x), sw - self._mw - 4)
        y = min(int(y), sh - self._mh - 4)
        self.geometry(f"{self._mw}x{self._mh}+{max(0, x)}+{max(0, y)}")
        bg = "white"
        try:
            self.attributes("-topmost", True)
            self.attributes("-transparentcolor", self.KEY)
            self.configure(bg=self.KEY)
            bg = self.KEY
        except tk.TclError:
            self.configure(bg="white")

        self.c = tk.Canvas(self, width=self._mw, height=self._mh, bg=bg,
                           highlightthickness=0, bd=0)
        self.c.pack(fill="both", expand=True)
        self._build()
        self.c.bind("<Motion>", self._on_motion)
        self.c.bind("<Button-1>", self._on_click)
        self.bind("<Escape>", lambda e: self.destroy())
        self.bind("<FocusOut>", lambda e: self.destroy())
        try:
            self.update_idletasks()
            self.lift()
            self.focus_force()
            self.grab_set()
        except tk.TclError:
            pass

    def _bounds(self):
        b, y = [], self._pad
        for i, (t, _cmd) in enumerate(self._items):
            if t == "-":
                y += self._sep_h
            else:
                b.append((i, y, y + self._row_h))
                y += self._row_h
        return b

    def _build(self):
        c = self.c
        c.delete("all")
        c.create_polygon(_round_rect_points(2, 2, self._mw - 2, self._mh - 2, sc(14)),
                         fill="white", outline="", smooth=True)
        y = self._pad
        for i, (t, _cmd) in enumerate(self._items):
            if t == "-":
                c.create_line(sc(16), y + self._sep_h // 2, self._mw - sc(16),
                              y + self._sep_h // 2, fill=HAIRLINE)
                y += self._sep_h
            else:
                fg = TEXT
                if i == self._hover:
                    c.create_polygon(
                        _round_rect_points(sc(6), y + sc(2), self._mw - sc(6),
                                           y + self._row_h - sc(2), sc(9)),
                        fill="#eaf7f4", outline="", smooth=True)
                    fg = ACCENT_ACTIVE
                c.create_text(sc(20), y + self._row_h // 2, anchor="w", text=t,
                              fill=fg, font=(FONT, 10))
                y += self._row_h

    def _on_motion(self, e):
        idx = -1
        for i, y0, y1 in self._bounds():
            if y0 <= e.y <= y1:
                idx = i
                break
        if idx != self._hover:
            self._hover = idx
            self.configure(cursor="hand2" if idx >= 0 else "")
            self._build()

    def _on_click(self, e):
        for i, y0, y1 in self._bounds():
            if y0 <= e.y <= y1:
                cmd = self._items[i][1]
                self.destroy()
                if cmd:
                    cmd()
                return
        self.destroy()


class App(tk.Tk):
    def __init__(self) -> None:
        super().__init__()
        init_ui_scale(self)   # 화면 DPI 로 UI 배율 계산(이후 sc()/위젯 크기에 반영)
        self.title(f"실시간 재고 현황(EcountERP) 및 평균 원가(Wizfasta) 비교  v{APP_VERSION}")
        self.geometry(f"{sc(1120)}x{sc(720)}")
        self.minsize(sc(940), sc(600))
        self.configure(bg=BG)

        # 로딩 화면(스플래시) — 메인은 준비될 때까지 숨김
        self.withdraw()
        self._splash = None
        self._splash_t0 = time.time()
        try:
            self._splash = Splash(self)
            self.update()                      # 스플래시 즉시 표시
            try:
                import pyi_splash               # PyInstaller 스플래시(있으면) 닫기
                pyi_splash.close()
            except Exception:                   # noqa: BLE001
                pass
        except Exception:                       # noqa: BLE001
            self._splash = None

        self._inventory_rows: list[dict] = []   # 탭1 조회 결과(원본: 가격비교 탭에서 사용)
        self._inventory_display: list[dict] = []  # 파싱 컬럼 포함 표시용(필터 적용 후)
        self._inventory_display_all: list[dict] = []  # 필터 전 전체(재필터용)
        self._inv_status_suffix: str = ""
        self._sort_col: str | None = None    # 정렬 기준 열(헤더 클릭)
        self._sort_desc: bool = False        # 내림차순 여부
        self._query_seq: int = 0             # 조회 시퀀스(백그라운드 가격조회 취소용)
        self._active_step: str | None = None   # 현재 진행 중 단계
        self._active_detail: str = ""
        self._active_since: float = 0.0
        self._anim_idx: int = 0
        self._anim_running: bool = False
        self._inventory_raw: dict | None = None
        self._compare_rows: list[dict] = []
        self._update_url: str = ""
        self._auth: dict | None = None   # 로그인 사용자 {username, role, status}

        # 인증/설정 변수 (메인 화면엔 표시하지 않고 '설정' 메뉴 다이얼로그에서 입력)
        self.var_com = tk.StringVar(value=FIXED_COM_CODE)     # 고정
        self.var_user = tk.StringVar(value=FIXED_USER_ID)     # 고정
        self.var_key = tk.StringVar(value=DEFAULT_API_CERT_KEY)  # 기본값(수정 가능)
        self.var_env = tk.StringVar(value="production")
        self.var_show_key = tk.BooleanVar(value=False)
        self.var_github = tk.StringVar(value=DEFAULT_GITHUB_REPO)
        self.var_model = tk.StringVar()      # 조회조건: 모델명 필터(클라이언트측)
        self.var_brand = tk.StringVar()      # 조회조건: 브랜드 필터(클라이언트측)
        self.var_wcorp = tk.StringVar()      # Wizfasta 업체코드
        self.var_wid = tk.StringVar()        # Wizfasta 아이디
        self.var_wpw = tk.StringVar()        # Wizfasta 비밀번호
        self.var_wshow = tk.BooleanVar(value=False)
        self.var_wkeep = tk.BooleanVar(value=False)   # 24시간 이후에도 계속 저장
        self._wiz_saved_at: float = 0.0      # Wizfasta 정보 저장 시각(만료 판정용)
        self.ent_key: ttk.Entry | None = None
        self.ent_wpw: ttk.Entry | None = None

        self._apply_theme()
        self._build_menu()
        self._build_header()

        # 탭 컨테이너 + 프레임(사용자 관리는 관리자 로그인 시에만 탭바에 노출)
        self._tabbar = None
        container = tk.Frame(self, bg=BG)
        container.pack(fill="both", expand=True, padx=14, pady=(0, 12))
        self._tab_container = container
        self.tab_inv = tk.Frame(container, bg=BG)
        self.tab_cmp = tk.Frame(container, bg=BG)
        self.tab_daily = tk.Frame(container, bg=BG)
        self.tab_useradmin = tk.Frame(container, bg=BG)
        self.tab_setup = tk.Frame(container, bg=BG)
        for f in (self.tab_inv, self.tab_cmp, self.tab_daily,
                  self.tab_useradmin, self.tab_setup):
            f.grid(row=0, column=0, sticky="nsew")
        container.rowconfigure(0, weight=1)
        container.columnconfigure(0, weight=1)
        self._tab_frames = {"inv": self.tab_inv, "cmp": self.tab_cmp,
                            "daily": self.tab_daily, "useradmin": self.tab_useradmin,
                            "setup": self.tab_setup}

        self._build_inventory_tab()
        self._build_compare_tab()
        self._build_daily_tab()
        self._build_user_admin_tab()
        self._build_setup_tab()
        self._build_tabbar(admin=False)   # 기본(비관리자) 탭바
        self._load_config()

        self._switch_tab("inv")   # 기본 탭

        # 로딩 화면에서 최신 버전 확인 → 있으면 묻지 않고 즉시 자동 설치, 없으면 메인 표시
        self.after(80, self._splash_update_then_start)

    def _build_tabbar(self, admin: bool) -> None:
        """탭바(재)구성. 관리자면 '일일현황' 옆에 '사용자 관리' 탭을 추가한다."""
        if getattr(self, "_tabbar", None) is not None:
            try:
                self._tabbar.destroy()
            except Exception:  # noqa: BLE001
                pass
        tabs = [("inv", "재고현황 조회", {}),
                ("cmp", "원가비교", {}),
                ("daily", "일일현황", {})]
        if admin:
            tabs.append(("useradmin", "사용자 관리", {}))
        tabs.append(("setup", "설치 현황",
                     {"side": "right", "height": 24, "font": (FONT, 9, "bold"),
                      "padx": 13, "radius": 12}))
        self._tabbar = RoundedTabBar(self, tabs, self._switch_tab, bg=BG)
        self._tabbar.pack(fill="x", padx=18, pady=(10, 4), before=self._tab_container)
        self._tabbar.select("inv")

    def _switch_tab(self, key: str) -> None:
        if key == "useradmin" and (self._auth or {}).get("role") != "admin":
            return
        frame = self._tab_frames.get(key)
        if frame is not None:
            frame.tkraise()
        if key == "daily":
            self._render_daily()
        elif key == "useradmin":
            self._render_user_admin()

    # ================= 디자인 테마 =================
    def _apply_theme(self) -> None:
        self.option_add("*Font", (FONT, 10))
        style = ttk.Style(self)
        try:
            style.theme_use("clam")
        except tk.TclError:
            pass
        style.configure(".", background=BG, foreground=TEXT, font=(FONT, 10))
        style.configure("TFrame", background=BG)
        style.configure("Tab.TFrame", background=BG)
        style.configure("Card.TFrame", background=CARD)
        style.configure("TLabel", background=BG, foreground=TEXT)
        style.configure("Muted.TLabel", background=BG, foreground=MUTED, font=(FONT, 9))
        style.configure("Status.TLabel", background=BG, foreground=ACCENT_ACTIVE, font=(FONT, 9, "bold"))
        # 입력 — 평평하고 얇은 헤어라인(모던)
        style.configure("TEntry", fieldbackground="white", bordercolor=BORDER,
                        lightcolor=BORDER, darkcolor=BORDER, relief="flat", padding=6)
        style.map("TEntry", bordercolor=[("focus", ACCENT)], lightcolor=[("focus", ACCENT)],
                  darkcolor=[("focus", ACCENT)])
        style.configure("TCombobox", fieldbackground="white", bordercolor=BORDER,
                        lightcolor=BORDER, darkcolor=BORDER, relief="flat", padding=6)
        # 카드형 LabelFrame — 얇은 헤어라인 + 깊은 제목
        style.configure("TLabelframe", background=BG, bordercolor=HAIRLINE,
                        relief="solid", borderwidth=1)
        style.configure("TLabelframe.Label", background=BG, foreground=ACCENT_ACTIVE,
                        font=(FONT, 10, "bold"))
        # 버튼 (대부분 RoundedButton 으로 대체)
        style.configure("TButton", background="#e5e7eb", foreground=TEXT,
                        bordercolor=BORDER, relief="flat", padding=(12, 7), font=(FONT, 10))
        style.map("TButton", background=[("active", "#d1d5db"), ("disabled", "#f0f1f3")],
                  foreground=[("disabled", "#9ca3af")])
        style.configure("TCheckbutton", background=BG, foreground=TEXT)
        # (탭은 RoundedTabBar 라운드 알약 컴포넌트로 대체 — ttk.Notebook 미사용)
        # 표(Treeview) — 넉넉한 행 높이·정돈된 헤더
        style.configure("Treeview", background="white", fieldbackground="white",
                        foreground=TEXT, rowheight=32, font=(FONT, 10), borderwidth=0)
        style.configure("Treeview.Heading", background=HEADING_BG, foreground=HEADING_FG,
                        font=(FONT, 10, "bold"), relief="flat", padding=8)
        style.map("Treeview.Heading", background=[("active", "#cdeee6")])
        style.map("Treeview", background=[("selected", SELECT_BG)],
                  foreground=[("selected", SELECT_FG)])
        # 스크롤바 — 얇고 모던하게
        style.configure("Vertical.TScrollbar", background="#d7deda", troughcolor=BG,
                        bordercolor=BG, arrowcolor=MUTED, relief="flat")
        style.configure("Horizontal.TScrollbar", background="#d7deda", troughcolor=BG,
                        bordercolor=BG, arrowcolor=MUTED, relief="flat")

    def _build_header(self) -> None:
        self._header_h = sc(66)
        self._header = tk.Canvas(self, height=self._header_h, highlightthickness=0, bd=0)
        self._header.pack(fill="x")
        self._header.bind("<Configure>", self._draw_header)

    def _draw_header(self, event=None) -> None:
        c = self._header
        w = (event.width if event else c.winfo_width()) or sc(1120)
        h = self._header_h
        c.delete("all")
        # 딥 틸 → 민트 그라데이션 + 하단 골드 헤어라인
        _paint_h_gradient(c, w, h, HEADER_DARK, ACCENT)
        c.create_rectangle(0, h - sc(3), w, h, fill=GOLD, outline="")
        # 좌측 골드 악센트 바
        c.create_rectangle(sc(20), h // 2 - sc(11), sc(24), h // 2 + sc(11), fill=GOLD, outline="")
        title = "실시간 재고 현황(EcountERP) 및 평균 원가(Wizfasta) 비교"
        c.create_text(sc(38), h // 2, anchor="w", text=title, fill="white",
                      font=(FONT, 15, "bold"))
        tw = tkfont.Font(family=FONT, size=15, weight="bold").measure(title)
        c.create_text(sc(38) + tw + sc(14), h // 2 + 1, anchor="w", text=f"v{APP_VERSION}",
                      fill=GOLD_SOFT, font=(FONT, 10))
        c.create_text(w - sc(20), h // 2, anchor="e",
                      text="설정 ▸ 인증 정보에서 API 키를 입력하세요",
                      fill=ACCENT_SOFT, font=(FONT, 9))

    # ================= 상단 메뉴 / 설정 =================
    def _build_menu(self) -> None:
        menubar = tk.Menu(self, background=ACCENT, foreground="white",
                          activebackground=ACCENT_ACTIVE, activeforeground="white")
        settings_menu = tk.Menu(menubar, tearoff=0, background="white", foreground=TEXT,
                                activebackground=ACCENT, activeforeground="white")
        settings_menu.add_command(label="인증 정보 설정...", command=self._open_settings)
        settings_menu.add_command(label="입고단가 캐시 비우기(갱신)", command=self._clear_price_cache)
        settings_menu.add_separator()
        settings_menu.add_command(label="종료", command=self.destroy)
        menubar.add_cascade(label="설정", menu=settings_menu)
        self._menubar = menubar
        self.config(menu=menubar)

    def _open_settings(self) -> None:
        win = tk.Toplevel(self)
        win.title("설정")
        win.transient(self)
        win.resizable(False, False)
        win.configure(bg=BG)
        pad = {"padx": 8, "pady": 5}

        auth = ttk.LabelFrame(win, text="인증 정보")
        auth.pack(fill="x", padx=12, pady=(12, 6))
        # 회사코드·사용자ID는 고정값(수정 불가) — readonly
        self.var_com.set(FIXED_COM_CODE)
        self.var_user.set(FIXED_USER_ID)
        ttk.Label(auth, text="회사코드").grid(row=0, column=0, sticky="e", **pad)
        ttk.Entry(auth, textvariable=self.var_com, width=30, state="readonly").grid(
            row=0, column=1, columnspan=2, sticky="w", **pad)
        ttk.Label(auth, text="사용자ID").grid(row=1, column=0, sticky="e", **pad)
        ttk.Entry(auth, textvariable=self.var_user, width=30, state="readonly").grid(
            row=1, column=1, columnspan=2, sticky="w", **pad)
        ttk.Label(auth, text="API 인증키").grid(row=2, column=0, sticky="e", **pad)
        self.ent_key = ttk.Entry(auth, textvariable=self.var_key, width=30, show="*")
        self.ent_key.grid(row=2, column=1, sticky="w", **pad)
        ttk.Checkbutton(auth, text="표시", variable=self.var_show_key,
                        command=self._toggle_key).grid(row=2, column=2, sticky="w", **pad)
        ttk.Label(auth, text="환경").grid(row=3, column=0, sticky="e", **pad)
        ttk.Combobox(auth, textvariable=self.var_env, width=28, state="readonly",
                     values=["production", "test"]).grid(row=3, column=1, columnspan=2, sticky="w", **pad)
        ttk.Label(auth, text="(production=운영 / test=상자테스트)", foreground="#666").grid(
            row=4, column=1, columnspan=2, sticky="w", padx=8)

        wiz = ttk.LabelFrame(win, text="Wizfasta 로그인 (가격비교용)")
        wiz.pack(fill="x", padx=12, pady=6)
        ttk.Label(wiz, text="업체코드").grid(row=0, column=0, sticky="e", **pad)
        ttk.Entry(wiz, textvariable=self.var_wcorp, width=30).grid(row=0, column=1, columnspan=2, sticky="w", **pad)
        ttk.Label(wiz, text="아이디").grid(row=1, column=0, sticky="e", **pad)
        ttk.Entry(wiz, textvariable=self.var_wid, width=30).grid(row=1, column=1, columnspan=2, sticky="w", **pad)
        ttk.Label(wiz, text="비밀번호").grid(row=2, column=0, sticky="e", **pad)
        self.ent_wpw = ttk.Entry(wiz, textvariable=self.var_wpw, width=30, show="*")
        self.ent_wpw.grid(row=2, column=1, sticky="w", **pad)
        ttk.Checkbutton(wiz, text="표시", variable=self.var_wshow,
                        command=self._toggle_wpw).grid(row=2, column=2, sticky="w", **pad)
        ttk.Checkbutton(wiz, text="24시간 이후에도 계속 저장",
                        variable=self.var_wkeep).grid(row=3, column=1, columnspan=2, sticky="w", **pad)
        ttk.Label(wiz, text="(체크 해제 시: 저장 후 24시간이 지나면 자동으로 공백 초기화)",
                  foreground="#666").grid(row=4, column=1, columnspan=2, sticky="w", padx=8)

        upd = ttk.LabelFrame(win, text="자동 업데이트")
        upd.pack(fill="x", padx=12, pady=6)
        ttk.Label(upd, text="GitHub 저장소").grid(row=0, column=0, sticky="e", **pad)
        ttk.Entry(upd, textvariable=self.var_github, width=30).grid(row=0, column=1, sticky="w", **pad)
        ttk.Label(upd, text="(예: Claudio-sys11/Price-Check)", foreground="#666").grid(
            row=1, column=1, sticky="w", padx=8)

        btns = ttk.Frame(win)
        btns.pack(fill="x", padx=12, pady=(6, 12))
        accent_button(btns, "저장", lambda: self._save_settings(win)).pack(side="right", padx=4)
        gray_button(btns, "닫기", win.destroy).pack(side="right", padx=4)

        win.grab_set()

    def _save_settings(self, win: tk.Toplevel) -> None:
        # Wizfasta 정보가 채워져 있으면 저장 시각 기록(24시간 만료 기준 갱신)
        if self.var_wcorp.get().strip() and self.var_wid.get().strip() and self.var_wpw.get():
            self._wiz_saved_at = time.time()
        else:
            self._wiz_saved_at = 0.0
        self._save_config()
        win.destroy()

    def _clear_price_cache(self) -> None:
        save_price_cache({})
        pmsg.showinfo("입고단가 캐시", "입고단가 캐시를 비웠습니다.\n다음 조회 때 품목등록에서 다시 받아옵니다.")

    # ================= 로그인 / 회원가입 / 사용자 관리 =================
    def _center_window(self, win, w: int, h: int) -> None:
        win.update_idletasks()
        sw, sh = win.winfo_screenwidth(), win.winfo_screenheight()
        x, y = (sw - w) // 2, (sh - h) // 3
        win.geometry(f"{w}x{h}+{max(0, x)}+{max(0, y)}")

    def _cursive_font(self, size, weight="bold"):
        """타이틀용 글꼴 — HY견명조 우선, 없으면 명조/궁서 계열, 최후 기본 폰트."""
        try:
            fams = set(tkfont.families())
        except Exception:   # noqa: BLE001
            fams = set()
        for name in ("HY견명조", "HY견명조체", "HY신명조", "궁서", "Batang"):
            if name in fams:
                return (name, size, weight)
        return (FONT, size, weight)

    def _script_font(self, size, weight="normal"):
        """영문 필기체(스크립트) 글꼴 — Segoe Script 우선, 없으면 다른 스크립트체."""
        try:
            fams = set(tkfont.families())
        except Exception:   # noqa: BLE001
            fams = set()
        for name in ("Segoe Script", "Lucida Handwriting", "Gabriola",
                     "Brush Script MT", "Ink Free"):
            if name in fams:
                return (name, size, weight)
        return (FONT, size, weight)

    def _show_login(self) -> None:
        KEY = "#FF00FE"   # 투명 처리 키 컬러(둥근 모서리)
        INDIGO, INDIGO_DK = "#1E0A5C", "#160848"   # 브랜드 인디고(럭셔리 포인트)
        dlg = tk.Toplevel(self)
        dlg.overrideredirect(True)
        dlg.protocol("WM_DELETE_WINDOW", self.destroy)   # 로그인 안 하면 종료
        w, h = sc(366), sc(524)
        sw, sh = dlg.winfo_screenwidth(), dlg.winfo_screenheight()
        gx, gy = (sw - w) // 2, (sh - h) // 3
        dlg.geometry(f"{w}x{h}+{max(0, gx)}+{max(0, gy)}")
        cbg = "white"
        try:
            dlg.attributes("-topmost", True)
            dlg.attributes("-transparentcolor", KEY)
            dlg.configure(bg=KEY)
            cbg = KEY
        except tk.TclError:
            dlg.configure(bg="white")

        c = tk.Canvas(dlg, width=w, height=h, bg=cbg, highlightthickness=0, bd=0)
        c.pack(fill="both", expand=True)
        m, rad = sc(7), sc(30)
        # 흰 카드 + 얇은 헤어라인 외곽(고급)
        c.create_polygon(_round_rect_points(m, m, w - m, h - m, rad),
                         fill="white", outline=HAIRLINE, width=1, smooth=True)
        # 상단 가운데 골드 미세 악센트
        c.create_polygon(
            _round_rect_points(w // 2 - sc(24), sc(18), w // 2 + sc(24), sc(22), sc(2)),
            fill=GOLD, outline="", smooth=True)

        # 중앙 브랜드 로고
        try:
            ic = tk.PhotoImage(file=resource_path("assets/app_icon.png"))
            f = max(1, round(ic.width() / sc(46)))
            ic = ic.subsample(f, f)
        except tk.TclError:
            ic = None
        self._login_icon = ic   # 참조 유지(가비지컬렉션 방지)
        if ic is not None:
            c.create_image(w // 2, sc(64), image=ic)

        # 필기체 워드마크 'Price Check' + 보조 명칭(2단계 작게) + 골드 짧은 구분선
        c.create_text(w // 2, sc(120), text="Price Check", fill=INDIGO,
                      font=self._script_font(26))
        c.create_text(w // 2, sc(148), text="원가비교 프로그램",
                      fill="#6b7280", font=self._cursive_font(9, "normal"))
        c.create_line(w // 2 - sc(28), sc(170), w // 2 + sc(28), sc(170),
                      fill=GOLD, width=1)

        # 하단 푸터(은은한 그레이)
        c.create_text(w // 2, h - sc(40), text="THE FEEL KOREA CO.,LTD.",
                      fill="#a3a8a6", font=(FONT, 10))
        c.create_text(w // 2, h - sc(22), text="Created by Claudio Lim",
                      fill="#bfc4c2", font=(FONT, 8))

        # 닫기(종료)
        cls = c.create_text(w - sc(26), sc(28), text="✕", fill="#c4cbc8", font=(FONT, 11))
        c.tag_bind(cls, "<Button-1>", lambda e: self.destroy())
        c.tag_bind(cls, "<Enter>", lambda e: c.itemconfig(cls, fill="#dc2626"))
        c.tag_bind(cls, "<Leave>", lambda e: c.itemconfig(cls, fill="#c4cbc8"))

        # 헤더 영역 드래그로 창 이동(테두리 없는 창)
        drag = {"x": 0, "y": 0}
        c.bind("<Button-1>", lambda e: drag.update(x=e.x, y=e.y))
        c.bind("<B1-Motion>", lambda e: dlg.geometry(
            f"+{dlg.winfo_x() + e.x - drag['x']}+{dlg.winfo_y() + e.y - drag['y']}"))

        # 입력 폼(중앙 정렬, 여백 넉넉)
        form = tk.Frame(dlg, bg="white")
        form.place(x=sc(36), y=sc(190), width=w - sc(72), height=h - sc(248))
        form.columnconfigure(0, weight=1)
        tk.Label(form, text="아이디", bg="white", fg="#9aa3a0",
                 font=(FONT, 9)).grid(row=0, column=0, sticky="w", pady=(0, 2))
        v_user = tk.StringVar()
        e_user = ttk.Entry(form, textvariable=v_user)
        e_user.grid(row=1, column=0, sticky="ew", ipady=sc(5))
        tk.Label(form, text="비밀번호", bg="white", fg="#9aa3a0",
                 font=(FONT, 9)).grid(row=2, column=0, sticky="w", pady=(12, 2))
        v_pw = tk.StringVar()
        e_pw = ttk.Entry(form, textvariable=v_pw, show="*")
        e_pw.grid(row=3, column=0, sticky="ew", ipady=sc(5))

        v_status = tk.StringVar(value="")
        tk.Label(form, textvariable=v_status, bg="white", fg="#dc2626",
                 font=(FONT, 9), wraplength=w - sc(90), justify="left").grid(
                     row=4, column=0, sticky="w", pady=(10, 2))

        btn_login = RoundedButton(form, "로그인", lambda: do_login(), bg="white",
                                  fill=INDIGO, fill_active=INDIGO_DK,
                                  fill_disabled="#b3aad4", fg="white",
                                  fg_disabled="#e7e2f5", height=42, radius=21)
        btn_login.grid(row=5, column=0, sticky="ew", pady=(8, 0))
        link = tk.Label(form, text="회원가입", bg="white", fg=INDIGO,
                        font=(FONT, 9, "underline"), cursor="hand2")
        link.grid(row=6, column=0, pady=(12, 0))
        link.bind("<Button-1>", lambda e: self._show_register(dlg))

        def set_busy(b):
            v_status.set("로그인 중…" if b else v_status.get())
            try:
                btn_login.configure(state="disabled" if b else "normal")
            except Exception:
                pass

        def do_login():
            u, p = v_user.get().strip(), v_pw.get()
            if not u or not p:
                v_status.set("아이디와 비밀번호를 입력하세요.")
                return
            set_busy(True)

            def work():
                try:
                    auth = backend.authenticate(u, p)
                    self.after(0, lambda: ok(auth))
                except backend.AuthError as e:
                    self.after(0, lambda: fail(str(e)))
                except backend.BackendError as e:
                    self.after(0, lambda: fail(f"서버 연결 오류: {e}"))
                except Exception as e:   # noqa: BLE001
                    self.after(0, lambda: fail(f"오류: {e}"))
            threading.Thread(target=work, daemon=True).start()

        def ok(auth):
            self._auth = auth
            try:
                dlg.grab_release()
            except Exception:
                pass
            dlg.destroy()
            self._apply_auth_role()
            self.deiconify()
            self.lift()
            self._render_daily()

        def fail(msg):
            v_status.set(msg)
            try:
                btn_login.configure(state="normal")
            except Exception:
                pass

        e_user.bind("<Return>", lambda e: e_pw.focus_set())
        e_pw.bind("<Return>", lambda e: do_login())
        # 메인 창이 withdraw 상태이므로 transient 로 묶지 않고 직접 표시
        dlg.lift()
        dlg.focus_force()
        dlg.after(350, lambda: dlg.winfo_exists() and dlg.attributes("-topmost", False))
        dlg.grab_set()
        e_user.focus_set()

    def _show_register(self, parent) -> None:
        if not backend.backend_enabled():
            pmsg.showwarning("회원가입 불가",
                             "공유 서버가 아직 설정되지 않아 회원가입을 사용할 수 없습니다.\n"
                             "관리자(THEFEELKOREA)로 로그인하거나 관리자에게 문의하세요.")
            return
        KEY = "#FF00FE"
        INDIGO, INDIGO_DK = "#1E0A5C", "#160848"
        dlg = tk.Toplevel(self)
        dlg.overrideredirect(True)
        w, h = sc(384), sc(600)
        sw, sh = dlg.winfo_screenwidth(), dlg.winfo_screenheight()
        gx, gy = (sw - w) // 2, (sh - h) // 4
        dlg.geometry(f"{w}x{h}+{max(0, gx)}+{max(0, gy)}")
        cbg = "white"
        try:
            dlg.attributes("-topmost", True)
            dlg.attributes("-transparentcolor", KEY)
            dlg.configure(bg=KEY)
            cbg = KEY
        except tk.TclError:
            dlg.configure(bg="white")

        c = tk.Canvas(dlg, width=w, height=h, bg=cbg, highlightthickness=0, bd=0)
        c.pack(fill="both", expand=True)
        m, rad = sc(7), sc(30)
        c.create_polygon(_round_rect_points(m, m, w - m, h - m, rad),
                         fill="white", outline=HAIRLINE, width=1, smooth=True)
        c.create_polygon(
            _round_rect_points(w // 2 - sc(24), sc(18), w // 2 + sc(24), sc(22), sc(2)),
            fill=GOLD, outline="", smooth=True)
        try:
            ic = tk.PhotoImage(file=resource_path("assets/app_icon.png"))
            f = max(1, round(ic.width() / sc(40)))
            ic = ic.subsample(f, f)
        except tk.TclError:
            ic = None
        self._reg_icon = ic   # 참조 유지
        if ic is not None:
            c.create_image(w // 2, sc(56), image=ic)
        c.create_text(w // 2, sc(104), text="Price Check", fill=INDIGO,
                      font=self._script_font(22))
        c.create_text(w // 2, sc(130), text="회원가입", fill="#6b7280", font=(FONT, 11))
        c.create_line(w // 2 - sc(28), sc(150), w // 2 + sc(28), sc(150),
                      fill=GOLD, width=1)
        c.create_text(w // 2, h - sc(40), text="THE FEEL KOREA CO.,LTD.",
                      fill="#a3a8a6", font=(FONT, 10))
        c.create_text(w // 2, h - sc(22), text="Created by Claudio Lim",
                      fill="#bfc4c2", font=(FONT, 8))

        def close_reg():
            try:
                parent.grab_set()
            except Exception:   # noqa: BLE001
                pass
            dlg.destroy()

        cls = c.create_text(w - sc(26), sc(28), text="✕", fill="#c4cbc8", font=(FONT, 11))
        c.tag_bind(cls, "<Button-1>", lambda e: close_reg())
        c.tag_bind(cls, "<Enter>", lambda e: c.itemconfig(cls, fill="#dc2626"))
        c.tag_bind(cls, "<Leave>", lambda e: c.itemconfig(cls, fill="#c4cbc8"))
        drag = {"x": 0, "y": 0}
        c.bind("<Button-1>", lambda e: drag.update(x=e.x, y=e.y))
        c.bind("<B1-Motion>", lambda e: dlg.geometry(
            f"+{dlg.winfo_x() + e.x - drag['x']}+{dlg.winfo_y() + e.y - drag['y']}"))

        form = tk.Frame(dlg, bg="white")
        form.place(x=sc(40), y=sc(166), width=w - sc(80), height=h - sc(220))
        form.columnconfigure(0, weight=1)
        v_u = tk.StringVar(); v_name = tk.StringVar()
        v_p = tk.StringVar(); v_p2 = tk.StringVar()
        fields = [("아이디 (3자 이상)", v_u, False),
                  ("사용자 이름", v_name, False),
                  ("비밀번호 (4자 이상)", v_p, True),
                  ("비밀번호 확인", v_p2, True)]
        entries = []
        for i, (lbl, var, hide) in enumerate(fields):
            tk.Label(form, text=lbl, bg="white", fg="#9aa3a0", font=(FONT, 9)).grid(
                row=i * 2, column=0, sticky="w", pady=(0 if i == 0 else 8, 2))
            e = ttk.Entry(form, textvariable=var, show="*" if hide else "")
            e.grid(row=i * 2 + 1, column=0, sticky="ew", ipady=sc(4))
            entries.append(e)

        v_status = tk.StringVar(value="")
        tk.Label(form, textvariable=v_status, bg="white", fg="#dc2626", font=(FONT, 9),
                 wraplength=w - sc(96), justify="left").grid(
                     row=8, column=0, sticky="w", pady=(8, 2))
        btn = RoundedButton(form, "가입 신청", lambda: submit(), bg="white",
                            fill=INDIGO, fill_active=INDIGO_DK, fill_disabled="#b3aad4",
                            fg="white", fg_disabled="#e7e2f5", height=42, radius=21)
        btn.grid(row=9, column=0, sticky="ew", pady=(6, 0))
        link = tk.Label(form, text="닫기", bg="white", fg="#6b7280",
                        font=(FONT, 9, "underline"), cursor="hand2")
        link.grid(row=10, column=0, pady=(10, 0))
        link.bind("<Button-1>", lambda e: close_reg())

        def submit():
            u, nm, p, p2 = (v_u.get().strip(), v_name.get().strip(),
                            v_p.get(), v_p2.get())
            if p != p2:
                v_status.set("비밀번호가 일치하지 않습니다.")
                return
            v_status.set("가입 신청 중…")
            btn.configure(state="disabled")

            def work():
                try:
                    backend.register(u, p, nm)
                    self.after(0, done)
                except (backend.AuthError, backend.BackendError) as e:
                    self.after(0, lambda: err(str(e)))
                except Exception as e:   # noqa: BLE001
                    self.after(0, lambda: err(f"오류: {e}"))

            def done():
                close_reg()
                pmsg.showinfo("가입 신청 완료",
                              "가입 신청이 접수되었습니다.\n관리자 승인 후 로그인할 수 있습니다.")
                try:
                    parent.grab_set()
                except Exception:   # noqa: BLE001
                    pass

            def err(msg):
                v_status.set(msg)
                btn.configure(state="normal")
            threading.Thread(target=work, daemon=True).start()

        entries[-1].bind("<Return>", lambda e: submit())
        dlg.lift()
        dlg.focus_force()
        dlg.after(350, lambda: dlg.winfo_exists() and dlg.attributes("-topmost", False))
        dlg.grab_set()
        entries[0].focus_set()

    def _apply_auth_role(self) -> None:
        """로그인 사용자에 맞춰 제목/탭을 갱신(관리자면 '사용자 관리' 탭 추가)."""
        a = self._auth or {}
        uname = a.get("username", "")
        role = a.get("role", "user")
        rolelbl = "관리자" if role == "admin" else "사용자"
        self.title(f"실시간 재고 현황(EcountERP) 및 평균 원가(Wizfasta) 비교  "
                   f"v{APP_VERSION}   [{uname} · {rolelbl}]")
        self._build_tabbar(admin=(role == "admin"))
        self._switch_tab("inv")

    # ----- 탭: 사용자 관리(관리자 전용) -----
    def _build_user_admin_tab(self) -> None:
        root = self.tab_useradmin
        ttk.Label(root, style="Muted.TLabel", justify="left",
                  text="회원가입 신청을 승인·거절하거나 사용자를 삭제합니다. (관리자 전용)\n"
                       "※ 목록에서 대상을 선택한 뒤 아래 버튼을 누르세요. 승인대기 사용자가 위에 표시됩니다."
                  ).pack(fill="x", padx=16, pady=(12, 2))
        btns = ttk.Frame(root)
        btns.pack(fill="x", padx=16, pady=(2, 6))
        accent_button(btns, "승인", lambda: self._useradmin_act("approve")).pack(side="left")
        gray_button(btns, "거절", lambda: self._useradmin_act("reject")).pack(side="left", padx=6)
        gray_button(btns, "삭제", lambda: self._useradmin_act("delete")).pack(side="left")
        gray_button(btns, "새로고침", self._render_user_admin).pack(side="left", padx=6)
        self.useradmin_status = tk.StringVar(value="")
        ttk.Label(btns, textvariable=self.useradmin_status, style="Status.TLabel").pack(side="right")

        tf = tk.Frame(root, bg=BORDER)
        tf.pack(fill="both", expand=True, padx=16, pady=(2, 14))
        cols = ("username", "name", "status", "role", "created_at")
        heads = {"username": "아이디", "name": "이름", "status": "상태",
                 "role": "권한", "created_at": "가입일시"}
        self.tree_useradmin = ttk.Treeview(tf, show="headings", columns=cols)
        for c in cols:
            self.tree_useradmin.heading(c, text=heads[c])
            self.tree_useradmin.column(c, anchor="center",
                                       width=(sc(170) if c == "username" else sc(120)),
                                       stretch=True)
        ysb = ttk.Scrollbar(tf, orient="vertical", command=self.tree_useradmin.yview)
        self.tree_useradmin.configure(yscrollcommand=ysb.set)
        self.tree_useradmin.grid(row=0, column=0, sticky="nsew", padx=1, pady=1)
        ysb.grid(row=0, column=1, sticky="ns")
        tf.rowconfigure(0, weight=1)
        tf.columnconfigure(0, weight=1)

    def _render_user_admin(self) -> None:
        if not hasattr(self, "tree_useradmin"):
            return
        if not backend.backend_enabled():
            self.useradmin_status.set("공유 서버가 설정되지 않았습니다(토큰 미설정).")
            return
        self.useradmin_status.set("불러오는 중…")
        STAT = {"pending": "승인대기", "approved": "승인됨", "rejected": "거절됨"}

        def fill(users):
            self.tree_useradmin.delete(*self.tree_useradmin.get_children())
            users.sort(key=lambda u: (u.get("status") != "pending", u.get("username", "")))
            for u in users:
                self.tree_useradmin.insert("", "end", iid=u.get("username", ""), values=(
                    u.get("username", ""), u.get("name", ""),
                    STAT.get(u.get("status"), u.get("status", "")),
                    "관리자" if u.get("role") == "admin" else "사용자",
                    u.get("created_at", "")))
            pend = sum(1 for u in users if u.get("status") == "pending")
            self.useradmin_status.set(f"총 {len(users)}명 · 승인대기 {pend}명")

        def work():
            try:
                users = backend.list_users()
                self.after(0, lambda: fill(users))
            except Exception as e:   # noqa: BLE001
                self.after(0, lambda: self.useradmin_status.set(f"오류: {e}"))
        threading.Thread(target=work, daemon=True).start()

    def _useradmin_act(self, action: str) -> None:
        if not hasattr(self, "tree_useradmin"):
            return
        sel = self.tree_useradmin.selection()
        if not sel:
            self.useradmin_status.set("대상 사용자를 선택하세요.")
            return
        uname = sel[0]
        labels = {"approve": "승인", "reject": "거절", "delete": "삭제"}
        self.useradmin_status.set(f"{labels.get(action, action)} 처리 중…")

        def work():
            try:
                if action == "approve":
                    backend.set_user_status(uname, "approved")
                elif action == "reject":
                    backend.set_user_status(uname, "rejected")
                else:
                    backend.delete_user(uname)
                self.after(0, self._render_user_admin)
            except Exception as e:   # noqa: BLE001
                self.after(0, lambda: self.useradmin_status.set(f"오류: {e}"))
        threading.Thread(target=work, daemon=True).start()

    # ================= 자동 업데이트 =================
    def _splash_update_then_start(self) -> None:
        """로딩 화면에서 최신 버전을 확인하고, 있으면 즉시 자동 설치한다."""
        if getattr(self, "_splash", None):
            self._splash.set_status("최신 버전 확인 중…")
        threading.Thread(target=self._splash_update_worker, daemon=True).start()

    def _splash_update_worker(self) -> None:
        cfg = {}
        try:
            if os.path.exists(CONFIG_PATH):
                with open(CONFIG_PATH, encoding="utf-8") as f:
                    cfg = json.load(f) or {}
        except (OSError, json.JSONDecodeError):
            cfg = {}
        if not (cfg.get("github_repo") or cfg.get("update_url")):
            cfg["github_repo"] = DEFAULT_GITHUB_REPO   # 설정 없어도 기본 저장소로 검사
        try:
            manifest = updater.check(cfg)
        except Exception:  # noqa: BLE001
            manifest = None

        def done():
            if manifest and manifest.get("url"):
                ver = manifest.get("version", "")
                if getattr(self, "_splash", None):
                    self._splash.set_status(f"새 버전 {ver} 자동 설치 중…")
                self._run_update(manifest["url"], ver)   # 묻지 않고 즉시 설치
            else:
                self._finish_splash()                    # 최신 → 메인 표시
        self.after(0, done)

    def _finish_splash(self) -> None:
        """로딩 화면을 닫고 메인 창을 보여준다(최소 표시시간 보장)."""
        elapsed = time.time() - getattr(self, "_splash_t0", 0)
        if elapsed < 1.1:   # 너무 빨리 깜빡이지 않도록 최소 표시
            self.after(int((1.1 - elapsed) * 1000), self._finish_splash)
            return
        sp = getattr(self, "_splash", None)
        if sp is not None:
            sp.close()
            self._splash = None
        if self._auth is None:
            self._show_login()   # 로그인 성공 시에만 메인 표시
            return
        self.deiconify()
        self.lift()

    def _set_update_status(self, msg: str) -> None:
        for setter in (
            lambda: self.status.set(msg),
            lambda: self.cmp_status.set(msg),
            lambda: self._splash.set_status(msg) if getattr(self, "_splash", None) else None,
        ):
            try:
                setter()
            except Exception:  # noqa: BLE001
                pass

    def _set_update_progress(self, pct: float) -> None:
        if getattr(self, "_splash", None):
            self._splash.set_progress(pct)

    def _run_update(self, url: str, ver: str = "") -> None:
        # 별도 업데이터 창(둥근 프리미엄 디자인)이 다운로드+설치를 한 창에서 진행한다.
        self._set_update_status(f"새 버전 {ver} 업데이트 준비 중…")
        target = sys.executable if getattr(sys, "frozen", False) else ""
        if spawn_install_updater(url, ver, target):
            self.after(300, self.destroy)   # 업데이터 창에 인계 → 본 앱 종료
            return

        # 폴백: 별도 창을 못 띄운 경우 기존 방식(본 창에서 다운로드 후 무인 설치)
        def prog(received, total):
            if total and total > 0:
                pct = received * 100.0 / total
                msg = (f"새 버전 {ver} 다운로드 중… {int(pct)}% "
                       f"({received / 1048576:.1f} / {total / 1048576:.1f} MB)")
                self.after(0, lambda p=pct, m=msg: (
                    self._set_update_progress(p), self._set_update_status(m)))

        def worker():
            try:
                path = updater.download_installer(url, progress=prog)
            except Exception:  # noqa: BLE001
                def fail():
                    if getattr(self, "_splash", None):
                        self._finish_splash()
                    else:
                        self._set_update_status("업데이트 확인 실패 — 다음 실행 시 재시도")
                self.after(0, fail)
                return

            def finish():
                self._set_update_progress(100)
                self._set_update_status("설치 중…")
                try:
                    updater.launch_installer(path, silent=True)
                finally:
                    self.after(800, self.destroy)
            self.after(0, finish)

        threading.Thread(target=worker, daemon=True).start()

    # ================= 탭1: 재고현황 =================
    def _build_inventory_tab(self) -> None:
        pad = {"padx": 6, "pady": 5}
        root = self.tab_inv

        ttk.Label(root, style="Muted.TLabel",
                  text="인증 정보는 상단 [설정] 메뉴 → '인증 정보 설정…' 에서 입력하세요."
                  ).pack(fill="x", padx=16, pady=(12, 2))

        cond = ttk.LabelFrame(root, text=" 조회 조건  (선택 — 비우면 전체) ")
        cond.pack(fill="x", padx=16, pady=6, ipady=4)
        self.var_base_date = tk.StringVar()
        self.var_prod = tk.StringVar()
        self.var_wh = tk.StringVar()
        # 조회 조건 — 한 줄 가로 배열
        cpad = {"padx": (8, 3), "pady": 6}
        epad = {"padx": (0, 10), "pady": 6}
        ttk.Label(cond, text="기준일자").grid(row=0, column=0, sticky="e", **cpad)
        ttk.Entry(cond, textvariable=self.var_base_date, width=12).grid(row=0, column=1, sticky="w", **epad)
        ttk.Label(cond, text="브랜드").grid(row=0, column=2, sticky="e", **cpad)
        ent_brand = ttk.Entry(cond, textvariable=self.var_brand, width=14)
        ent_brand.grid(row=0, column=3, sticky="w", **epad)
        ttk.Label(cond, text="모델명").grid(row=0, column=4, sticky="e", **cpad)
        ent_model = ttk.Entry(cond, textvariable=self.var_model, width=14)
        ent_model.grid(row=0, column=5, sticky="w", **epad)
        ttk.Label(cond, text="품목코드").grid(row=0, column=6, sticky="e", **cpad)
        ttk.Entry(cond, textvariable=self.var_prod, width=12).grid(row=0, column=7, sticky="w", **epad)
        ttk.Label(cond, text="창고코드").grid(row=0, column=8, sticky="e", **cpad)
        ttk.Entry(cond, textvariable=self.var_wh, width=12).grid(row=0, column=9, sticky="w", **epad)
        # 조회 후 브랜드/모델명을 바꾸면 재조회 없이 즉시 재필터
        ent_brand.bind("<KeyRelease>", self._on_filter_change)
        ent_model.bind("<KeyRelease>", self._on_filter_change)
        self.btn_reset = gray_button(cond, "↺ 조건 초기화", self._on_reset_conditions)
        self.btn_reset.grid(row=0, column=10, sticky="w", padx=(6, 8), pady=6)
        ttk.Label(cond, style="Muted.TLabel",
                  text="※ 기준일자는 YYYYMMDD. 브랜드·모델명은 조회 결과를 부분일치로 필터링합니다. 결과는 브랜드 순으로 정렬됩니다.").grid(
                      row=1, column=0, columnspan=11, sticky="w", padx=8, pady=(2, 0))

        btns = ttk.Frame(root)
        btns.pack(fill="x", padx=16, pady=(2, 6))
        self.btn_query = accent_button(btns, "🔍  재고현황 조회", self._on_query)
        self.btn_query.pack(side="left")
        self.btn_inv_csv = gray_button(
            btns, "Excel 내보내기",
            lambda: export_rows_excel(self._inventory_display, "재고현황.xlsx"))
        self.btn_inv_csv.configure(state="disabled")
        self.btn_inv_csv.pack(side="left", padx=8)
        self.btn_sub_csv = gray_button(btns, "소계/평균만 내보내기", self._export_subtotals)
        self.btn_sub_csv.configure(state="disabled")
        self.btn_sub_csv.pack(side="left")
        self.btn_clear_cache = gray_button(btns, "🗑  캐시 비우기", self._clear_price_cache)
        self.btn_clear_cache.pack(side="left", padx=8)
        self.status = tk.StringVar(value="대기 중")
        ttk.Label(btns, textvariable=self.status, style="Status.TLabel").pack(side="right")

        tf = tk.Frame(root, bg=BORDER)   # 1px 테두리 느낌의 카드
        tf.pack(fill="both", expand=True, padx=16, pady=(2, 14))
        self.tree_inv = ttk.Treeview(tf, show="headings")
        ysb = ttk.Scrollbar(tf, orient="vertical", command=self.tree_inv.yview)
        xsb = ttk.Scrollbar(tf, orient="horizontal", command=self.tree_inv.xview)
        self.tree_inv.configure(yscrollcommand=ysb.set, xscrollcommand=xsb.set)
        self.tree_inv.grid(row=0, column=0, sticky="nsew", padx=1, pady=1)
        ysb.grid(row=0, column=1, sticky="ns")
        xsb.grid(row=1, column=0, sticky="ew")
        tf.rowconfigure(0, weight=1)
        tf.columnconfigure(0, weight=1)
        self._enable_tree_copy(self.tree_inv, self.status)

    def _toggle_key(self) -> None:
        if self.ent_key is not None:
            self.ent_key.configure(show="" if self.var_show_key.get() else "*")

    def _toggle_wpw(self) -> None:
        if self.ent_wpw is not None:
            self.ent_wpw.configure(show="" if self.var_wshow.get() else "*")

    # ================= 탭2: 가격비교 =================
    def _build_compare_tab(self) -> None:
        root = self.tab_cmp
        self._wiz_rows: list[dict] = []

        # 상단: 좌(안내·버튼·칩) / 우(진행 상황 — 우측 상단, 내용에 맞춘 크기)
        top = ttk.Frame(root)
        top.pack(fill="x", padx=16, pady=(12, 2))

        chk = ttk.LabelFrame(top, text=" 진행 상황 ")
        chk.pack(side="right", anchor="n", padx=(14, 0))
        self._step_labels = {}
        for key, label in WIZ_STEPS:
            lb = ttk.Label(chk, text=f"⬜  {label}", style="Muted.TLabel")
            lb.pack(anchor="w", padx=10, pady=1)
            self._step_labels[key] = (lb, label)

        left = ttk.Frame(top)
        left.pack(side="left", fill="both", expand=True)

        ttk.Label(left, style="Muted.TLabel", justify="left",
                  text="① 재고현황 탭에서 조회 후 → ② [Wizfasta 원가 가져오기]로 상품DB 원가를 받아 모델명으로 비교합니다.\n"
                       "EcountERP 값은 재고현황 소계의 '평균원가'(모델별 가중평균)를 사용합니다. "
                       "(60초 초과 시 자동 재시작 / [중단]으로 멈춤)\n"
                       "※ 셀을 클릭/드래그해 원하는 범위를 선택하고 Ctrl+C로 복사하세요. "
                       "모델명 셀을 더블클릭하면 재고현황 탭에서 해당 모델로 조회합니다."
                  ).pack(anchor="w", pady=(0, 4))

        btns = ttk.Frame(left)
        btns.pack(fill="x", pady=(2, 4))
        self.btn_wiz = accent_button(btns, "🛒  Wizfasta 원가 가져오기", self._on_fetch_wizfasta)
        self.btn_wiz.pack(side="left")
        self.btn_wiz_stop = RoundedButton(
            btns, "■ 중단", self._on_stop_wizfasta, bg=BG,
            fill=DANGER, fill_active=DANGER_ACTIVE, fill_disabled=DANGER_DISABLED,
            fg="white", fg_disabled="#fbeaea")
        self.btn_wiz_stop.configure(state="disabled")
        self.btn_wiz_stop.pack(side="left", padx=8)
        self.btn_cmp_csv = gray_button(
            btns, "비교결과 Excel 내보내기",
            lambda: export_rows_excel(self._compare_rows, "가격비교.xlsx"))
        self.btn_cmp_csv.configure(state="disabled")
        self.btn_cmp_csv.pack(side="left", padx=8)
        self.cmp_status = tk.StringVar(
            value="먼저 재고현황 탭에서 조회한 뒤, Wizfasta 원가 가져오기를 누르세요."
        )
        ttk.Label(btns, textvariable=self.cmp_status, style="Status.TLabel").pack(side="right")

        # 결과 요약(둥근 통계 칩) — 좌측 영역 안에서 우측 정렬
        self.cmp_summary = tk.Frame(left, bg=BG)
        self.cmp_summary.pack(fill="x", pady=(2, 0))
        chip_box = tk.Frame(self.cmp_summary, bg=BG)
        chip_box.pack(side="right")
        self._cmp_cat = None   # 대시보드 분류 필터(None=전체)
        cw = 122
        self.chip_total = StatChip(chip_box, "전체 (Wiz)", fill="#e0f7f1", fg="#0f766e",
                                   cw=cw, command=lambda: self._set_cmp_category(None))
        self.chip_diff = StatChip(chip_box, "원가차이", fill=DIFF_BG, fg=DIFF_FG,
                                  cw=cw, command=lambda: self._set_cmp_category("diff"))
        self.chip_nostock = StatChip(chip_box, "미입고", fill="#fdeccb", fg="#8a5a0a",
                                     cw=cw, command=lambda: self._set_cmp_category("nostock"))
        self.chip_unmatch = StatChip(chip_box, "미매칭", fill=UNMATCH_BG, fg="#4b5563",
                                     cw=cw, command=lambda: self._set_cmp_category("unmatched"))
        self.chip_same = StatChip(chip_box, "원가일치", fill="#d7f5ed", fg="#0d9488",
                                  cw=cw, command=lambda: self._set_cmp_category("same"))
        for c in (self.chip_total, self.chip_diff, self.chip_nostock,
                  self.chip_unmatch, self.chip_same):
            c.pack(side="left", padx=(8, 0))

        # 검색 필터(브랜드·모델명) — 비교 결과를 부분일치로 즉시 필터
        self.var_cmp_brand = tk.StringVar()
        self.var_cmp_model = tk.StringVar()
        filt = ttk.Frame(root)
        filt.pack(fill="x", padx=16, pady=(8, 0))
        ttk.Label(filt, text="🔎  검색", style="Status.TLabel").pack(side="left", padx=(0, 10))
        ttk.Label(filt, text="브랜드").pack(side="left")
        ent_cb = ttk.Entry(filt, textvariable=self.var_cmp_brand, width=18)
        ent_cb.pack(side="left", padx=(4, 14))
        ttk.Label(filt, text="모델명").pack(side="left")
        ent_cm = ttk.Entry(filt, textvariable=self.var_cmp_model, width=22)
        ent_cm.pack(side="left", padx=(4, 14))
        gray_button(filt, "필터 지우기", self._clear_cmp_filter, height=30, radius=15, padx=14).pack(side="left")
        ent_cb.bind("<KeyRelease>", self._on_cmp_filter_change)
        ent_cm.bind("<KeyRelease>", self._on_cmp_filter_change)

        # 원가비교 표 — 셀 단위 선택/드래그 범위 선택/복사가 가능한 시트(tksheet)
        tf = tk.Frame(root, bg=BORDER)
        tf.pack(fill="both", expand=True, padx=16, pady=(6, 14))
        self._cmp_headers = []
        self.sheet_cmp = Sheet(
            tf, headers=[], data=[],
            theme="light blue",
            header_bg=HEADING_BG, header_fg=HEADING_FG,
            header_border_fg=HAIRLINE, table_grid_fg="#eef2f1",
            font=(FONT, 10, "normal"), header_font=(FONT, 10, "bold"),
            table_selected_cells_bg=SELECT_BG, table_selected_cells_fg=SELECT_FG,
            table_selected_box_cells_fg=GOLD,
            table_selected_rows_bg=SELECT_BG, table_selected_rows_fg=SELECT_FG,
            table_selected_columns_bg=SELECT_BG, table_selected_columns_fg=SELECT_FG,
            show_x_scrollbar=True, show_y_scrollbar=True)
        # 셀 선택 + 드래그 범위 선택 + 복사 (편집은 비활성)
        self.sheet_cmp.enable_bindings(
            "single_select", "drag_select", "ctrl_select", "shift_select",
            "select_all", "copy", "arrowkeys",
            "column_width_resize", "double_click_column_resize",
            "right_click_popup_menu", "rc_select")
        self.sheet_cmp.pack(fill="both", expand=True, padx=1, pady=1)
        self._cmp_sort_col = None
        self._cmp_sort_desc = False
        # 모델명 셀 더블클릭 → 재고현황 조회 / 머리글 클릭 → 오름·내림 정렬 (캔버스 직접 바인딩)
        self.sheet_cmp.MT.bind("<Double-Button-1>", self._on_cmp_cell_double, add="+")
        self.sheet_cmp.CH.bind("<Button-1>", self._on_cmp_header_click, add="+")
        # 깔끔한 스타일: 행 번호 숨김, 적절한 행/머리글 높이
        try:
            self.sheet_cmp.hide("row_index")
        except Exception:  # noqa: BLE001
            pass
        # 데이터가 셀 안에서 세로 가운데로 보이도록 행 높이를 '텍스트높이+여백'으로 맞춤
        # (tksheet 는 텍스트를 셀 상단에 그리므로, 위/아래 여백이 같아지게 행 높이를 조정)
        try:
            th = int(self.sheet_cmp.MT.table_txt_height)
        except Exception:  # noqa: BLE001
            th = 17
        self._cmp_row_h = max(sc(24), th + sc(6))
        try:
            self.sheet_cmp.set_options(default_row_height=self._cmp_row_h,
                                       default_header_height="1")
        except Exception:  # noqa: BLE001
            pass
        # 선택 박스(점선 테두리) 색상을 그레이로 (기본 검정 → 회색)
        try:
            self.sheet_cmp.set_options(
                table_selected_box_cells_fg="#9ca3af",
                table_selected_box_rows_fg="#9ca3af",
                table_selected_box_columns_fg="#9ca3af")
        except Exception:  # noqa: BLE001
            pass

    def _on_cmp_cell_double(self, event=None) -> None:
        """원가비교 시트 셀 더블클릭: 모델명 셀이면 재고현황 탭으로 이동·조회."""
        try:
            r = self.sheet_cmp.identify_row(event)
            c = self.sheet_cmp.identify_column(event)
        except Exception:  # noqa: BLE001
            return
        if r is None or c is None:
            return
        headers = self._cmp_headers
        if 0 <= c < len(headers) and headers[c] == "모델명":
            try:
                model = str(self.sheet_cmp.get_cell_data(r, c) or "")
            except Exception:  # noqa: BLE001
                model = ""
            if model:
                self._goto_inventory_for_model(model)

    def _on_cmp_header_click(self, event=None) -> None:
        """머리글 클릭: 해당 열 오름차순, 같은 열 다시 클릭하면 내림차순."""
        try:
            c = self.sheet_cmp.identify_column(event)
        except Exception:  # noqa: BLE001
            return
        headers = self._cmp_headers
        if c is None or not (0 <= c < len(headers)):
            return
        col = headers[c]
        if self._cmp_sort_col == col:
            self._cmp_sort_desc = not self._cmp_sort_desc
        else:
            self._cmp_sort_col = col
            self._cmp_sort_desc = False
        if getattr(self, "_compare_rows_all", []):
            self._render_compare()

    def _fill_compare_sheet(self, rows: list[dict]) -> None:
        """원가비교 시트를 채우고 행 색상·열정렬·열폭을 적용한다."""
        headers = []
        for r in rows:
            for k in r.keys():
                if not k.startswith("_") and k not in headers:
                    headers.append(k)
        if not headers:
            headers = ["브랜드", "모델명", "파스타원가", "평균원가(ERP)", "차이",
                       "파스타재고", "실재고(ERP)", "재고차이", "매칭", "비고"]
        self._cmp_headers = headers
        data = [[r.get(h, "") for h in headers] for r in rows]

        sh = self.sheet_cmp
        # 정렬 중인 열에는 ▲/▼ 표시(표시용 머리글 — 식별용 _cmp_headers 는 원래 이름 유지)
        # 주의: 지역변수명에 sc 사용 금지 — 모듈 함수 sc()(DPI 배율) 를 가려 열폭 계산이 깨진다.
        sort_col = getattr(self, "_cmp_sort_col", None)
        arrow = " ▼" if getattr(self, "_cmp_sort_desc", False) else " ▲"
        disp_headers = [(h + arrow) if h == sort_col else h for h in headers]
        sh.headers(disp_headers)
        sh.set_sheet_data(data, reset_col_positions=True, reset_row_positions=True, redraw=False)
        try:   # 모든 행 높이를 동일하게(세로 가운데 정렬 효과)
            sh.set_all_row_heights(getattr(self, "_cmp_row_h", 24), redraw=False)
        except Exception:  # noqa: BLE001
            pass
        try:
            sh.dehighlight_all()
        except Exception:  # noqa: BLE001
            pass
        for i, r in enumerate(rows):
            tag = r.get("_tag")
            if tag == "diff":
                sh.highlight_rows([i], bg=DIFF_BG, fg=DIFF_FG, redraw=False)
            elif tag == "nostock":
                sh.highlight_rows([i], bg="#fff4e0", fg="#8a5a0a", redraw=False)
            elif tag == "unmatched":
                sh.highlight_rows([i], bg=UNMATCH_BG, fg=UNMATCH_FG, redraw=False)
        # 금액·수량 컬럼 우측정렬, 매칭 컬럼 가운데 정렬
        money_cols = [ci for ci, h in enumerate(headers) if h in MONEY_COLS]
        if money_cols:
            try:
                sh.align_columns(columns=money_cols, align="e", redraw=False)
            except Exception:  # noqa: BLE001
                pass
        if "매칭" in headers:
            try:
                sh.align_columns(columns=[headers.index("매칭")], align="center", redraw=False)
            except Exception:  # noqa: BLE001
                pass
        # 각 칸을 머리글·내용 중 넓은 쪽에 맞춰 자동 폭 조정(전부 보이게)
        fnt = tkfont.Font(family=FONT, size=10)
        fnt_b = tkfont.Font(family=FONT, size=10, weight="bold")
        for ci, h in enumerate(headers):
            w = fnt_b.measure(str(h))
            for r in rows:
                w = max(w, fnt.measure(str(r.get(h, ""))))
            w = min(max(w + sc(34), sc(64)), sc(520))   # 여백 + 최소/최대 제한
            try:
                sh.column_width(column=ci, width=w, redraw=False)
            except Exception:  # noqa: BLE001
                pass
        sh.redraw()

    def _goto_inventory_for_model(self, model: str) -> None:
        """재고현황 조회 탭으로 전환하고 해당 모델명으로 필터(조회)한다."""
        self._tabbar.select("inv")
        self._switch_tab("inv")
        self.var_brand.set("")
        self.var_model.set(model)
        if getattr(self, "_inventory_display_all", []):
            self._render_inventory()
            self.status.set(f"'{model}' 모델로 필터했습니다.")
        else:
            pmsg.showinfo(
                "재고현황 조회 필요",
                f"'{model}' (으)로 보려면 먼저 [재고현황 조회]를 눌러 조회하세요.\n"
                "조회하면 이 모델명 필터가 자동 적용됩니다.")

    # ---- 표 복사 기능(우클릭 메뉴/드래그 선택/Ctrl+C/더블클릭) ----
    def _enable_tree_copy(self, tree: ttk.Treeview, status_var: tk.StringVar) -> None:
        tree.configure(selectmode="extended")
        tree._copy_status = status_var          # 복사 결과 안내용
        tree._ctx_cell = (None, None)
        tree._drag_anchor = None

        def on_right(e):
            row = tree.identify_row(e.y)
            col = tree.identify_column(e.x)
            if not row:
                return
            if row not in tree.selection():
                tree.selection_set(row)
            tree._ctx_cell = (row, col)
            items = [
                ("이 셀 복사", lambda: self._copy_cell(tree)),
                ("선택 행 복사", lambda: self._copy_rows(tree)),
                ("-", None),
                ("전체 복사 (머리글 포함)",
                 lambda: self._copy_rows(tree, all_rows=True, header=True)),
            ]
            PremiumMenu(self, items, e.x_root, e.y_root)

        # 드래그 선택(누른 행 → 끌어간 행 범위 선택)
        def on_press(e):
            tree._drag_anchor = tree.identify_row(e.y)

        def on_drag(e):
            anchor = getattr(tree, "_drag_anchor", None)
            cur = tree.identify_row(e.y)
            if not anchor or not cur:
                return
            items = tree.get_children()
            try:
                i1, i2 = items.index(anchor), items.index(cur)
            except ValueError:
                return
            lo, hi = (i1, i2) if i1 <= i2 else (i2, i1)
            tree.selection_set(items[lo:hi + 1])
            tree.see(cur)

        tree.bind("<Button-3>", on_right)
        tree.bind("<Button-1>", on_press, add="+")
        tree.bind("<B1-Motion>", on_drag, add="+")
        tree.bind("<Control-c>", lambda e: self._copy_rows(tree))
        tree.bind("<Control-C>", lambda e: self._copy_rows(tree))
        tree.bind("<Double-1>", lambda e: self._copy_cell_at(tree, e))

    def _set_copy_status(self, tree, msg: str) -> None:
        sv = getattr(tree, "_copy_status", None)
        if sv is not None:
            sv.set(msg)

    def _copy_text(self, text: str) -> None:
        self.clipboard_clear()
        self.clipboard_append(text)

    def _copy_cell_at(self, tree, e) -> None:
        row = tree.identify_row(e.y)
        col = tree.identify_column(e.x)
        if row:
            tree._ctx_cell = (row, col)
            self._copy_cell(tree)

    def _copy_cell(self, tree) -> None:
        row, col = getattr(tree, "_ctx_cell", (None, None))
        cols = list(tree["columns"])
        if not row or not col or not cols:
            return
        try:
            idx = int(str(col).replace("#", "")) - 1
        except ValueError:
            return
        if idx < 0 or idx >= len(cols):
            return
        val = str(tree.set(row, cols[idx]))
        self._copy_text(val)
        self._set_copy_status(tree, f"복사됨: {val[:40]}")

    def _copy_rows(self, tree, all_rows: bool = False, header: bool = False) -> None:
        items = tree.get_children() if all_rows else tree.selection()
        cols = list(tree["columns"])
        if not items or not cols:
            return
        lines = []
        if header:
            lines.append("\t".join(cols))
        for it in items:
            lines.append("\t".join(str(tree.set(it, c)) for c in cols))
        self._copy_text("\n".join(lines))
        self._set_copy_status(tree, f"{len(items)}행 복사됨 (클립보드)")

    def _reset_steps(self) -> None:
        self._active_step = None
        for key, (lb, label) in self._step_labels.items():
            lb.configure(text=f"⬜  {label}", foreground=MUTED)

    def _set_step(self, key: str, detail: str = "") -> None:
        """현재 진행 단계를 지정한다. 이전 단계는 ✅, 현재 단계는 ⏳(애니메이션)."""
        order = [k for k, _ in WIZ_STEPS]
        if key not in order:
            return
        idx = order.index(key)
        if key != self._active_step:
            self._active_since = time.time()
        self._active_step = key
        self._active_detail = detail
        for i, (k, label) in enumerate(WIZ_STEPS):
            lb = self._step_labels[k][0]
            if i < idx:
                lb.configure(text=f"✅  {label}", foreground="#0d9488")
            elif i > idx:
                lb.configure(text=f"⬜  {label}", foreground=MUTED)
        self._render_active_step()

    def _render_active_step(self) -> None:
        """현재 단계를 [아이콘] [이름] · [세부정보]  [스피너] [경과초] 로 실시간 표시."""
        if not self._active_step:
            return
        lb, label = self._step_labels[self._active_step]
        spin = SPINNER[self._anim_idx % len(SPINNER)]
        elapsed = int(time.time() - self._active_since)
        det = f"  ·  {self._active_detail}" if self._active_detail else ""
        lb.configure(text=f"⏳  {label}{det}   {spin} {elapsed}s", foreground=ACCENT)

    def _start_anim(self) -> None:
        self._anim_running = True
        self._anim_idx = 0
        self._tick_anim()

    def _tick_anim(self) -> None:
        if not self._anim_running:
            return
        self._anim_idx += 1
        self._render_active_step()
        self.after(150, self._tick_anim)

    def _stop_anim(self) -> None:
        self._anim_running = False

    def _mark_all_done(self) -> None:
        self._stop_anim()
        self._active_step = None
        for key, (lb, label) in self._step_labels.items():
            lb.configure(text=f"✅  {label}", foreground="#0d9488")

    def _on_fetch_wizfasta(self) -> None:
        if not getattr(self, "_inventory_display_all", []):
            pmsg.showwarning(
                "재고 먼저 조회",
                "먼저 ① 재고현황 탭에서 조회해 주세요.\n(EcountERP 모델명·입고단가가 있어야 비교됩니다.)")
            return
        corp, uid, pw = self.var_wcorp.get().strip(), self.var_wid.get().strip(), self.var_wpw.get()
        if not (corp and uid and pw):
            pmsg.showwarning(
                "Wizfasta 로그인 정보 필요",
                "[설정] → 'Wizfasta 로그인'에 업체코드·아이디·비밀번호를 입력해 주세요.")
            return
        self._reset_steps()
        self._wiz_stop = threading.Event()
        self.btn_wiz.configure(state="disabled")
        self.btn_wiz_stop.configure(state="normal")
        self.btn_cmp_csv.configure(state="disabled")
        self.cmp_status.set("진행 중…")
        self._set_step("start", "Chrome 준비")
        self._start_anim()   # 실시간 활동 표시(블로킹 단계에서도 멈춘 듯 보이지 않게)
        threading.Thread(target=self._do_fetch_wizfasta, args=(corp, uid, pw), daemon=True).start()

    def _on_stop_wizfasta(self) -> None:
        if getattr(self, "_wiz_stop", None):
            self._wiz_stop.set()
        self.cmp_status.set("중단 요청 중…")
        self.btn_wiz_stop.configure(state="disabled")

    def _do_fetch_wizfasta(self, corp: str, uid: str, pw: str) -> None:
        try:
            import wizfasta_selenium
        except Exception as exc:  # noqa: BLE001
            self.after(0, lambda: self._wiz_failed(f"Selenium 모듈 로드 실패: {exc}"))
            return
        prog = lambda key, detail="": self.after(0, lambda k=key, d=detail: self._set_step(k, d))
        max_attempts = 3
        for attempt in range(1, max_attempts + 1):
            if self._wiz_stop.is_set():
                self.after(0, self._wiz_cancelled)
                return
            try:
                rows = wizfasta_selenium.fetch_wizfasta_costs(
                    corp, uid, pw, progress=prog, headless=True,
                    start_timeout=60, overall_timeout=60,
                    should_stop=lambda: self._wiz_stop.is_set())
            except wizfasta_selenium.FetchStopped:
                self.after(0, self._wiz_cancelled)
                return
            except (wizfasta_selenium.FetchTimeout, TimeoutError):
                # 60초 초과 → 재시작
                self.after(0, lambda a=attempt: (self._reset_steps(),
                                                 self.cmp_status.set(f"60초 초과 — 재시작 {a}/{max_attempts}…"),
                                                 self._set_step("start", "재시작")))
                continue
            except Exception as exc:  # noqa: BLE001
                self.after(0, lambda e=exc: self._wiz_failed(f"Wizfasta 가져오기 실패: {e}"))
                return
            self.after(0, lambda r=rows: self._wiz_done(r))
            return
        self.after(0, lambda: self._wiz_failed("60초 내 완료되지 않아 재시작을 반복했습니다. 잠시 후 다시 시도하세요."))

    def _wiz_cancelled(self) -> None:
        self._stop_anim()
        if self._active_step:
            lb, label = self._step_labels[self._active_step]
            lb.configure(text=f"■  {label} (중단됨)", foreground="#6b7280")
        self.btn_wiz.configure(state="normal")
        self.btn_wiz_stop.configure(state="disabled")
        self.cmp_status.set("중단됨 — 다시 실행할 수 있습니다.")

    def _wiz_failed(self, msg: str) -> None:
        self._stop_anim()
        if self._active_step:   # 멈춘 단계를 ❌로 표시
            lb, label = self._step_labels[self._active_step]
            lb.configure(text=f"❌  {label}", foreground="#dc2626")
        self.btn_wiz.configure(state="normal")
        self.btn_wiz_stop.configure(state="disabled")
        self.cmp_status.set("실패")
        pmsg.showerror("Wizfasta 가져오기 실패", msg)

    def _wiz_done(self, wiz_rows: list[dict]) -> None:
        self.btn_wiz.configure(state="normal")
        self.btn_wiz_stop.configure(state="disabled")
        self._wiz_rows = wiz_rows or []
        if not self._wiz_rows:
            self._stop_anim()
            self.cmp_status.set("Wizfasta 원가 0건 — 로그인/조회 상태를 확인하세요.")
            pmsg.showwarning("데이터 없음", "Wizfasta에서 원가를 받지 못했습니다(0건).")
            return
        self._set_step("match")
        ecount_data = [d for d in self._inventory_display_all]  # 모델명·입고단가 포함
        self._compare_rows_all = cmp.build_cost_comparison(self._wiz_rows, ecount_data)
        self.btn_cmp_csv.configure(state="normal")
        self._mark_all_done()
        self._record_daily_status()   # 오늘 날짜로 대시보드 집계 스냅샷 저장
        self._render_compare()   # 현재 필터(브랜드/모델명) 적용해 표시

    def _render_compare(self) -> None:
        """비교 결과에 브랜드/모델명 필터(부분일치)를 적용해 표·칩·상태를 갱신."""
        allrows = getattr(self, "_compare_rows_all", [])
        bf = self.var_cmp_brand.get().strip().lower()
        mf = self.var_cmp_model.get().strip().lower()
        mf_norm = cmp.normalize_model(mf) if mf else ""

        def keep(r):
            if bf and bf not in str(r.get("브랜드", "")).lower():
                return False
            if mf:
                model = str(r.get("모델명", ""))
                if mf in model.lower():
                    return True
                if mf_norm:
                    mn = cmp.normalize_model(model)
                    if mn and (mf_norm in mn or mn in mf_norm):
                        return True
                return False
            return True

        rows_bm = [r for r in allrows if keep(r)]   # 브랜드/모델명 필터 적용

        # 칩 수치(분류별)는 브랜드/모델명 필터 기준으로 계산
        total = len(rows_bm)
        n_diff = sum(1 for r in rows_bm if r.get("_tag") == "diff")
        n_nostock = sum(1 for r in rows_bm if r.get("_tag") == "nostock")
        n_unmatch = sum(1 for r in rows_bm if r.get("_tag") == "unmatched")
        n_same = sum(1 for r in rows_bm if r.get("_tag") == "same")
        self.chip_total.set_value(f"{total:,}")
        self.chip_diff.set_value(f"{n_diff:,}")
        self.chip_nostock.set_value(f"{n_nostock:,}")
        self.chip_unmatch.set_value(f"{n_unmatch:,}")
        self.chip_same.set_value(f"{n_same:,}")
        # 전체 대비 비율(%) — 분류 칩에 표시
        def _pct(n):
            return f"{round(n * 100 / total)}%" if total else "0%"
        self.chip_diff.set_pct(_pct(n_diff))
        self.chip_nostock.set_pct(_pct(n_nostock))
        self.chip_unmatch.set_pct(_pct(n_unmatch))
        self.chip_same.set_pct(_pct(n_same))
        # 선택된 분류 칩 강조
        self.chip_total.set_active(self._cmp_cat is None)
        self.chip_diff.set_active(self._cmp_cat == "diff")
        self.chip_nostock.set_active(self._cmp_cat == "nostock")
        self.chip_unmatch.set_active(self._cmp_cat == "unmatched")
        self.chip_same.set_active(self._cmp_cat == "same")

        # 대시보드 분류 필터 적용
        rows = ([r for r in rows_bm if r.get("_tag") == self._cmp_cat]
                if self._cmp_cat else list(rows_bm))

        # 정렬: 머리글 클릭 시 해당 열 / 기본은 브랜드 → 모델명 오름차순
        sc = getattr(self, "_cmp_sort_col", None)
        if sc:
            rows.sort(key=lambda r: _sort_key(r.get(sc, "")),
                      reverse=getattr(self, "_cmp_sort_desc", False))
        else:
            rows.sort(key=lambda r: (str(r.get("브랜드", "")).upper(),
                                     str(r.get("모델명", "")).upper()))

        self._compare_rows = rows
        self._fill_compare_sheet(rows)

        full = len(allrows)
        catname = {"diff": "원가차이", "nostock": "미입고",
                   "unmatched": "미매칭", "same": "원가일치"}.get(self._cmp_cat)
        if catname:
            self.cmp_status.set(f"[{catname}] {len(rows):,}건 / 전체 {full:,}건")
        elif total == full:
            self.cmp_status.set(f"완료 — 총 {total:,}건")
        else:
            self.cmp_status.set(f"필터 {total:,}건 / 전체 {full:,}건")

    def _set_cmp_category(self, cat) -> None:
        """대시보드 칩 클릭 → 해당 분류로 필터(같은 칩 다시 누르면 전체로 토글)."""
        self._cmp_cat = None if (cat is not None and self._cmp_cat == cat) else cat
        if getattr(self, "_compare_rows_all", []):
            self._render_compare()

    def _on_cmp_filter_change(self, event=None) -> None:
        if getattr(self, "_compare_rows_all", []):
            self._render_compare()

    def _clear_cmp_filter(self) -> None:
        self.var_cmp_brand.set("")
        self.var_cmp_model.set("")
        self._cmp_cat = None
        self._cmp_sort_col = None      # 기본 정렬(브랜드→모델명 오름차순)로 복귀
        self._cmp_sort_desc = False
        if getattr(self, "_compare_rows_all", []):
            self._render_compare()

    # ================= 탭3: 설치 현황 =================
    def _build_setup_tab(self) -> None:
        root = self.tab_setup
        ttk.Label(root, text="이 프로그램 사용에 필요한 항목",
                  font=(FONT, 13, "bold"), background=BG, foreground=TEXT
                  ).pack(anchor="w", padx=18, pady=(16, 2))
        ttk.Label(root, style="Muted.TLabel",
                  text="설치 상태를 확인하고, 없으면 오른쪽 버튼(링크)으로 설치하세요."
                  ).pack(anchor="w", padx=18, pady=(0, 10))

        self._setup_items = [
            {"name": "Google Chrome", "desc": "Wizfasta 원가 가져오기에 필요 (필수)",
             "check": self._check_chrome, "btn": "Chrome 설치/다운로드",
             "url": "https://www.google.com/chrome/"},
            {"name": "chromedriver", "desc": "Chrome 자동 제어 드라이버",
             "check": self._check_driver, "btn": None, "url": None},
            {"name": f"이 프로그램 (v{APP_VERSION})", "desc": "최신 버전 확인 / 다운로드",
             "check": self._check_app, "btn": "최신 릴리스 페이지",
             "url": "https://github.com/Claudio-sys11/Price-Check/releases/latest"},
        ]
        self._setup_status = {}
        box = ttk.Frame(root)
        box.pack(fill="x", padx=18)
        for it in self._setup_items:
            row = ttk.LabelFrame(box, text=f" {it['name']} ")
            row.pack(fill="x", pady=5, ipady=3)
            ttk.Label(row, style="Muted.TLabel", text=it["desc"]).grid(
                row=0, column=0, sticky="w", padx=10, pady=2)
            st = ttk.Label(row, text="확인 중…", foreground=MUTED)
            st.grid(row=1, column=0, sticky="w", padx=10, pady=2)
            self._setup_status[it["name"]] = st
            if it["btn"]:
                gray_button(row, it["btn"],
                            lambda u=it["url"]: self._open_url(u)).grid(
                    row=0, column=1, rowspan=2, sticky="e", padx=10, pady=6)
            row.columnconfigure(0, weight=1)

        accent_button(root, "🔄  다시 확인", self._check_requirements).pack(
            anchor="w", padx=18, pady=12)
        self.after(300, self._check_requirements)

    def _open_url(self, url: str | None) -> None:
        import webbrowser
        if url:
            webbrowser.open(url)

    def _check_chrome(self):
        try:
            from wizfasta_selenium import find_chrome
            p = find_chrome()
        except Exception:
            p = None
        if p:
            return True, f"설치됨: {p}"
        return False, "미설치 — 오른쪽 [Chrome 설치/다운로드] 로 설치하세요"

    def _check_driver(self):
        return True, "앱이 자동으로 설치·관리합니다 (별도 설치 불필요)"

    def _check_app(self):
        return True, f"현재 v{APP_VERSION} — 자동 업데이트 지원"

    def _check_requirements(self) -> None:
        for it in self._setup_items:
            try:
                ok, detail = it["check"]()
            except Exception as exc:  # noqa: BLE001
                ok, detail = False, f"확인 실패: {exc}"
            st = self._setup_status[it["name"]]
            icon = "✅" if ok else "❌"
            st.configure(text=f"{icon}  {detail}",
                         foreground=("#0d9488" if ok else "#dc2626"))

    # ================= 설정 =================
    def _load_config(self) -> None:
        if not os.path.exists(CONFIG_PATH):
            return
        try:
            with open(CONFIG_PATH, encoding="utf-8") as f:
                cfg = json.load(f)
        except (OSError, json.JSONDecodeError):
            return
        # 회사코드·사용자ID는 고정값(저장값 무시), API 인증키는 저장값 우선·없으면 기본값
        self.var_com.set(FIXED_COM_CODE)
        self.var_user.set(FIXED_USER_ID)
        self.var_key.set(cfg.get("API_CERT_KEY") or DEFAULT_API_CERT_KEY)
        self.var_env.set(cfg.get("ENV", "production"))
        self._update_url = cfg.get("update_url", "")
        self.var_github.set(cfg.get("github_repo", "") or DEFAULT_GITHUB_REPO)
        # Wizfasta 로그인: 최초 공백 / '계속 저장' 아니면 24시간 후 공백 초기화
        wz = cfg.get("wizfasta") or {}
        keep = bool(wz.get("keep", False))
        saved_at = float(wz.get("saved_at", 0) or 0)
        self.var_wkeep.set(keep)
        expired = (not keep) and saved_at and (time.time() - saved_at > 86400)
        if wz.get("corp") and not expired:
            self.var_wcorp.set(wz.get("corp", ""))
            self.var_wid.set(wz.get("id", ""))
            self.var_wpw.set(wz.get("pw", ""))
            self._wiz_saved_at = saved_at
        else:
            self.var_wcorp.set("")
            self.var_wid.set("")
            self.var_wpw.set("")
            self._wiz_saved_at = 0.0
            if expired:   # 만료분은 디스크에서도 공백 처리
                cfg["wizfasta"] = {"corp": "", "id": "", "pw": "", "saved_at": 0, "keep": False}
                try:
                    with open(CONFIG_PATH, "w", encoding="utf-8") as f:
                        json.dump(cfg, f, ensure_ascii=False, indent=2)
                except OSError:
                    pass
        payload = (cfg.get("inventory") or {}).get("payload", {})
        self.var_base_date.set(payload.get("BASE_DATE", ""))
        self.var_prod.set(payload.get("PROD_CD", ""))
        self.var_wh.set(payload.get("WH_CD", ""))

    def _current_config(self) -> dict:
        return {
            "COM_CODE": FIXED_COM_CODE,          # 고정
            "USER_ID": FIXED_USER_ID,            # 고정
            "API_CERT_KEY": (self.var_key.get().strip() or DEFAULT_API_CERT_KEY),
            "LAN_TYPE": "ko-KR",
            "ENV": self.var_env.get(),
            "update_url": getattr(self, "_update_url", ""),
            "github_repo": self.var_github.get().strip(),
            "wizfasta": {
                "corp": self.var_wcorp.get().strip(),
                "id": self.var_wid.get().strip(),
                "pw": self.var_wpw.get(),
                "saved_at": getattr(self, "_wiz_saved_at", 0.0),
                "keep": self.var_wkeep.get(),
            },
            "inventory": {
                "endpoint": "/OAPI/V2/InventoryBalance/GetListInventoryBalanceStatus",
                "payload": {
                    "BASE_DATE": self.var_base_date.get().strip(),
                    "PROD_CD": self.var_prod.get().strip(),
                    "WH_CD": self.var_wh.get().strip(),
                },
            },
        }

    def _save_config(self) -> None:
        try:
            with open(CONFIG_PATH, "w", encoding="utf-8") as f:
                json.dump(self._current_config(), f, ensure_ascii=False, indent=2)
        except OSError as exc:
            pmsg.showerror("저장 실패", f"설정 저장 중 오류:\n{exc}")
            return
        pmsg.showinfo("저장 완료", f"설정을 저장했습니다.\n{CONFIG_PATH}")

    # ================= 조회 실행 =================
    def _on_query(self) -> None:
        cfg = self._current_config()
        if not cfg["COM_CODE"] or not cfg["USER_ID"] or not cfg["API_CERT_KEY"]:
            pmsg.showwarning("입력 필요", "회사코드 / 사용자ID / API 인증키를 모두 입력하세요.")
            return
        self._query_seq = getattr(self, "_query_seq", 0) + 1  # 새 조회 → 이전 백그라운드 가격조회 중단
        self.btn_query.configure(state="disabled")
        self.btn_inv_csv.configure(state="disabled")
        self.status.set("조회 중...")
        try:
            with open(CONFIG_PATH, "w", encoding="utf-8") as f:
                json.dump(cfg, f, ensure_ascii=False, indent=2)
        except OSError:
            pass
        threading.Thread(target=self._do_query, args=(cfg, self._query_seq), daemon=True).start()

    # 재고현황 엔드포인트: 풍부한 ByLocation(창고/품목명/규격 포함) 우선, 권한 없으면 기본으로 폴백
    RICH_EP = "/OAPI/V2/InventoryBalance/GetListInventoryBalanceStatusByLocation"
    BASIC_EP = "/OAPI/V2/InventoryBalance/GetListInventoryBalanceStatus"

    def _do_query(self, cfg: dict, my_seq: int = 0) -> None:
        from datetime import date
        try:
            client = EcountClient(
                com_code=cfg["COM_CODE"], user_id=cfg["USER_ID"],
                api_cert_key=cfg["API_CERT_KEY"], lan_type=cfg.get("LAN_TYPE", "ko-KR"),
                env=cfg.get("ENV", "production"),
            )
            payload = {k: v for k, v in cfg["inventory"]["payload"].items() if v}
            if not payload.get("BASE_DATE"):
                payload["BASE_DATE"] = date.today().strftime("%Y%m%d")
            client.get_zone()
            client.login()  # 한 번만 로그인 (이후 엔드포인트 재시도에 세션 재사용)
        except EcountApiError as exc:
            self.after(0, self._query_failed, str(exc))
            return
        except Exception as exc:  # noqa: BLE001
            self.after(0, self._query_failed, f"예기치 못한 오류: {exc}")
            return

        note = ""
        try:
            # 1순위: 창고별 재고현황 (창고코드/창고명/품목명/재고 포함)
            data = client.get_inventory(endpoint=self.RICH_EP, payload=payload)
            rows = cmp.extract_ecount_rows(data)
        except EcountApiError as exc:
            # 권한 없으면 기본 재고현황으로 대체
            note = f"창고별 재고(ByLocation) 사용 불가 → 기본 재고로 대체. 사유: {exc}"
            try:
                data = client.get_inventory(endpoint=self.BASIC_EP, payload=payload)
                rows = cmp.extract_ecount_rows(data)
            except EcountApiError as exc2:
                self.after(0, self._query_failed, str(exc2))
                return

        # 입고단가: 품목코드별로 품목등록(IN_PRICE)을 개별 조회해 매칭(캐시 활용)
        code_f = cmp.detect_ecount_fields(rows).get("품목코드") or "PROD_CD"
        inv_codes = list(dict.fromkeys(
            str(r.get(code_f, "")).strip() for r in rows if str(r.get(code_f, "")).strip()))
        cache = load_price_cache()

        # 1) 우선 캐시에 있는 입고단가로 즉시 표시
        price_map = {c: cache[c] for c in inv_codes if c in cache}
        all_display = cmp.build_inventory_display(rows, price_map=price_map)
        self.after(0, self._query_done, data, rows, all_display, note)

        # 2) 누락된 품목코드 입고단가를 백그라운드로 받아 채우고 자동 갱신
        missing = [c for c in inv_codes if c not in cache]
        if missing:
            def prog(d, t):
                self.after(0, lambda d=d, t=t: self.status.set(
                    f"입고단가 매칭 중… {d}/{t} (완료 후 자동 갱신)"))
            try:
                fetched = client.get_prices(
                    missing, progress=prog,
                    should_stop=lambda: getattr(self, "_query_seq", 0) != my_seq)
            except EcountApiError:
                fetched = {}
            if fetched and getattr(self, "_query_seq", 0) == my_seq:
                cache.update(fetched)
                save_price_cache(cache)
                full_map = {c: cache[c] for c in inv_codes if c in cache}
                new_all = cmp.build_inventory_display(rows, price_map=full_map)
                self.after(0, self._prices_updated, rows, new_all, len(fetched))

    def _query_failed(self, msg: str) -> None:
        self.btn_query.configure(state="normal")
        self.status.set("실패")
        pmsg.showerror("조회 실패", msg)

    def _query_done(self, data: dict, rows: list[dict],
                    all_display: list[dict], product_note: str = "") -> None:
        self._inventory_raw = data
        self._inventory_rows = rows                  # 원본(가격비교 탭에서 사용)
        self._inventory_display_all = all_display    # 필터 전 전체(재필터용)
        self.btn_query.configure(state="normal")
        if not rows:
            self._inventory_display = []
            fill_tree(self.tree_inv, [])
            self.status.set("완료 — 표시할 행 없음")
            self.btn_inv_csv.configure(state="normal")
            pmsg.showinfo(
                "조회 완료",
                "응답을 받았지만 재고 행을 자동으로 찾지 못했습니다.\n"
                "'CSV 내보내기'로 원문 구조를 확인하세요.",
            )
            return

        # 연동 항목 안내(상태바 접두)
        has_name = any(d.get("브랜드") for d in all_display)
        has_price = any(
            str(d.get("입고단가", "")).replace(",", "").lstrip("-") not in ("", "0")
            for d in all_display
        )
        has_wh = any("창고코드" in d for d in all_display)
        if has_name or has_price:
            parts = []
            if has_name:
                parts.append("품목명 분해" + (" + 창고별" if has_wh else ""))
            if has_price:
                parts.append("입고단가/총단가")
            self._inv_status_suffix = f"({', '.join(parts)} 적용, 소계 포함)"
        else:
            self._inv_status_suffix = "(품목명/입고단가 미적용: 창고별/품목 조회 API 권한 필요)"
            if product_note:
                pmsg.showwarning(
                    "재고 항목 일부 미연동",
                    "재고수량은 조회됐지만 브랜드/모델명/사이즈/입고일자·창고 정보를 채울 "
                    "품목명을 가져오지 못했습니다.\n\n"
                    f"사유: {product_note}\n\n"
                    "EcountERP에서 '창고별 재고현황(GetListInventoryBalanceStatusByLocation)' "
                    "또는 '품목 조회(GetBasicProductsList)' API 권한을 켜면 다음 조회부터 "
                    "자동으로 모든 항목이 채워집니다.",
                )

        self.btn_inv_csv.configure(state="normal")
        self.btn_sub_csv.configure(state="normal")
        self._render_inventory(initial=True)

    def _render_inventory(self, initial: bool = False) -> None:
        """이미 받아온 전체 데이터에 현재 브랜드/모델명 필터 + 브랜드 정렬 + 소계를 적용해 표시."""
        allrows = getattr(self, "_inventory_display_all", [])
        bf = self.var_brand.get().strip().lower()
        mf = self.var_model.get().strip().lower()
        mf_norm = cmp.normalize_model(mf) if mf else ""   # 정규화 매칭(가격비교 모델명 대응)

        def _match(d):
            if bf and bf not in str(d.get("브랜드", "")).lower():
                return False
            if mf:
                model = str(d.get("모델명", ""))
                if mf in model.lower():
                    return True
                if mf_norm:
                    mn = cmp.normalize_model(model)
                    # 양방향 부분일치(가격비교 모델명에 브랜드 접두가 붙는 경우 대응)
                    if mn and (mf_norm in mn or mn in mf_norm):
                        return True
                return False
            return True

        filtered = [d for d in allrows if _match(d)]
        if self._sort_col:   # 헤더 클릭 정렬(오름/내림)
            col = self._sort_col
            if col == "사이즈":
                keyf = lambda d: _size_key(d.get("사이즈", ""))
            else:
                keyf = lambda d: _sort_key(d.get(col, ""))
            filtered.sort(key=keyf, reverse=self._sort_desc)
        else:                # 기본 정렬: 브랜드 → 모델명 → 사이즈(S,M,L/숫자) → 입고일자
            filtered.sort(key=lambda d: (
                str(d.get("브랜드", "")),
                str(d.get("모델명", "")),
                _size_key(d.get("사이즈", "")),
                str(d.get("입고일자", "")),
                str(d.get("품목코드", "")),
                str(d.get("창고코드", "")),
            ))
        rows = cmp.add_subtotals(filtered)
        # 브랜드/모델명으로 검색(필터) 중이면, 검색 결과 전체의 합계·평균 행을 맨 아래에 추가
        if (bf or mf) and filtered:
            rows = rows + [self._grand_total_row(filtered)]
        self._inventory_display = rows
        fill_tree(self.tree_inv, self._inventory_display,
                  sort_callback=self._on_sort_column,
                  sort_col=self._sort_col, sort_desc=self._sort_desc)

        total = len(allrows)
        shown = len(filtered)
        suffix = getattr(self, "_inv_status_suffix", "")
        if shown == total:
            self.status.set(f"완료 — {total}건 {suffix}".rstrip())
        else:
            self.status.set(f"필터 {shown}건 / 전체 {total}건 {suffix}".rstrip())

    def _grand_total_row(self, filtered: list[dict]) -> dict:
        """검색(필터) 결과 전체의 합계·평균을 담은 행을 만든다."""
        cols = [k for k in filtered[0].keys() if not k.startswith("_")]
        sum_qty = sum(cmp._to_number(r.get("재고수량")) for r in filtered)
        sum_total = sum(cmp._to_number(r.get("총단가")) for r in filtered)
        if sum_qty:
            avg = sum_total / sum_qty
        else:
            prices = [cmp._to_number(r.get("입고단가")) for r in filtered]
            avg = (sum(prices) / len(prices)) if prices else 0
        g = {k: "" for k in cols}
        if "브랜드" in g:
            g["브랜드"] = "■ 검색 합계/평균"
        label_col = "창고명" if "창고명" in g else "사이즈"
        if label_col in g:
            g[label_col] = f"{len(filtered)}건"
        g["재고수량"] = int(sum_qty)
        g["입고단가"] = f"{int(round(avg)):,}"     # 평균단가
        g["총단가"] = f"{int(round(sum_total)):,}"
        g["_subtotal"] = True
        g["_grand"] = True
        return g

    def _on_filter_change(self, event=None) -> None:
        """조회조건(브랜드/모델명) 변경 시: 이미 받아온 데이터에서 즉시 재필터(재조회 없음)."""
        if getattr(self, "_inventory_display_all", []):
            self._render_inventory()

    def _on_reset_conditions(self) -> None:
        """조회 조건(기준일자·브랜드·모델명·품목코드·창고코드)을 모두 비우고 즉시 재필터."""
        for var in (self.var_base_date, self.var_brand, self.var_model,
                    self.var_prod, self.var_wh):
            var.set("")
        if getattr(self, "_inventory_display_all", []):
            self._render_inventory()
        try:
            self.status.set("조회 조건을 초기화했습니다.")
        except Exception:
            pass

    def _on_sort_column(self, col: str) -> None:
        """열 머리글 클릭: 같은 열이면 오름↔내림 토글, 다른 열이면 그 열 오름차순."""
        if not getattr(self, "_inventory_display_all", []):
            return
        if self._sort_col == col:
            self._sort_desc = not self._sort_desc
        else:
            self._sort_col = col
            self._sort_desc = False
        self._render_inventory()

    def _export_subtotals(self) -> None:
        """현재 표의 소계/평균 행만 (브랜드·모델명 포함) 엑셀로 내보낸다."""
        subs = [d for d in self._inventory_display if d.get("_subtotal")]
        if not subs:
            pmsg.showwarning("데이터 없음", "내보낼 소계가 없습니다. 먼저 재고현황을 조회하세요.")
            return
        rows = [{
            "브랜드": d.get("브랜드", ""),
            "모델명": d.get("모델명", ""),
            "재고수량": d.get("재고수량", ""),
            "평균단가": d.get("입고단가", ""),
            "총단가": d.get("총단가", ""),
        } for d in subs]
        export_rows_excel(rows, "소계평균.xlsx")

    # ================= 탭: 일일현황 =================
    def _record_daily_status(self) -> None:
        """원가비교 완료 시 오늘 날짜로 대시보드 집계를 저장(같은 날은 최신으로 갱신)."""
        rows = getattr(self, "_compare_rows_all", [])
        if not rows:
            return
        total = len(rows)
        rec = {
            "date": time.strftime("%Y-%m-%d"),
            "time": time.strftime("%H:%M"),
            "total": total,
            "diff": sum(1 for r in rows if r.get("_tag") == "diff"),
            "nostock": sum(1 for r in rows if r.get("_tag") == "nostock"),
            "unmatched": sum(1 for r in rows if r.get("_tag") == "unmatched"),
            "same": sum(1 for r in rows if r.get("_tag") == "same"),
        }
        # 로컬 백업 저장(오프라인 대비)
        hist = [h for h in load_daily_status() if h.get("date") != rec["date"]]
        hist.append(rec)
        hist.sort(key=lambda h: h.get("date", ""))
        save_daily_status(hist)
        # 공유 저장소에 기록(모두가 볼 수 있도록)
        if backend.backend_enabled():
            def work():
                try:
                    backend.record_daily(rec)
                except Exception:   # noqa: BLE001
                    pass
                self.after(0, self._render_daily)
            threading.Thread(target=work, daemon=True).start()
        elif hasattr(self, "tree_daily"):
            self._render_daily()

    def _build_daily_tab(self) -> None:
        root = self.tab_daily
        ttk.Label(root, style="Muted.TLabel", justify="left",
                  text="원가비교를 실행할 때마다 그날의 대시보드 집계(전체·원가차이·미입고·미매칭·원가일치)가 "
                       "자동 저장되어 일자별 추이를 확인할 수 있습니다.\n"
                       "※ 같은 날 여러 번 비교하면 가장 최근 결과로 갱신됩니다. 괄호 안은 전체 대비 비율(%)입니다."
                  ).pack(fill="x", padx=16, pady=(12, 2))

        btns = ttk.Frame(root)
        btns.pack(fill="x", padx=16, pady=(2, 6))
        self.btn_daily_refresh = accent_button(btns, "↻  새로고침", self._render_daily)
        self.btn_daily_refresh.pack(side="left")
        self.btn_daily_csv = gray_button(btns, "Excel 내보내기", self._export_daily)
        self.btn_daily_csv.pack(side="left", padx=8)
        self.daily_status = tk.StringVar(value="")
        ttk.Label(btns, textvariable=self.daily_status, style="Status.TLabel").pack(side="right")

        # 원가차이 추이 그래프 (막대) — 일자별
        cf = tk.Frame(root, bg=BORDER)
        cf.pack(fill="x", padx=16, pady=(2, 8))
        self.daily_chart = tk.Canvas(cf, bg=CARD, height=sc(210),
                                     highlightthickness=0, bd=0)
        self.daily_chart.pack(fill="x", expand=True, padx=1, pady=1)
        self.daily_chart.bind("<Configure>", lambda e: self._draw_diff_chart())

        tf = tk.Frame(root, bg=BORDER)
        tf.pack(fill="both", expand=True, padx=16, pady=(2, 14))
        cols = ("date", "time", "total", "diff", "nostock", "unmatched", "same")
        heads = {"date": "일자", "time": "갱신시각", "total": "전체",
                 "diff": "원가차이", "nostock": "미입고",
                 "unmatched": "미매칭", "same": "원가일치"}
        self.tree_daily = ttk.Treeview(tf, show="headings", columns=cols)
        for c in cols:
            self.tree_daily.heading(c, text=heads[c])
            self.tree_daily.column(
                c, anchor="center",
                width=(sc(130) if c == "date" else sc(108)), stretch=True)
        ysb = ttk.Scrollbar(tf, orient="vertical", command=self.tree_daily.yview)
        self.tree_daily.configure(yscrollcommand=ysb.set)
        self.tree_daily.grid(row=0, column=0, sticky="nsew", padx=1, pady=1)
        ysb.grid(row=0, column=1, sticky="ns")
        tf.rowconfigure(0, weight=1)
        tf.columnconfigure(0, weight=1)
        self._enable_tree_copy(self.tree_daily, self.daily_status)

    def _render_daily(self) -> None:
        if not hasattr(self, "tree_daily"):
            return
        if backend.backend_enabled():
            self.daily_status.set("공유 기록 불러오는 중…")
            def work():
                try:
                    hist = backend.load_daily()
                except Exception:   # noqa: BLE001
                    hist = load_daily_status()   # 오프라인 폴백(로컬 백업)
                self.after(0, lambda: self._fill_daily(hist))
            threading.Thread(target=work, daemon=True).start()
        else:
            self._fill_daily(load_daily_status())

    def _fill_daily(self, hist: list) -> None:
        if not hasattr(self, "tree_daily"):
            return
        hist = list(hist or [])
        hist.sort(key=lambda h: (h.get("date", ""), h.get("time", "")), reverse=True)
        self._daily_cache = hist
        self.tree_daily.delete(*self.tree_daily.get_children())

        def cell(n, total):
            return f"{n:,} ({round(n * 100 / total)}%)" if total else f"{n:,}"

        for h in hist:
            t = h.get("total", 0)
            self.tree_daily.insert("", "end", values=(
                h.get("date", ""), h.get("time", ""), f"{t:,}",
                cell(h.get("diff", 0), t), cell(h.get("nostock", 0), t),
                cell(h.get("unmatched", 0), t), cell(h.get("same", 0), t)))
        shared = " (공유)" if backend.backend_enabled() else ""
        self.daily_status.set(
            f"총 {len(hist)}일치 기록{shared}" if hist
            else "기록 없음 — 원가비교를 실행하면 자동 저장됩니다.")
        self._draw_diff_chart()

    def _draw_diff_chart(self) -> None:
        """일자별 '원가차이' 수치를 막대그래프로 그린다(과거→현재, 최근 24일)."""
        cv = getattr(self, "daily_chart", None)
        if cv is None:
            return
        cv.delete("all")
        W = cv.winfo_width()
        H = cv.winfo_height()
        if W < 60 or H < 60:
            return
        cv.create_text(sc(14), sc(13), text="원가차이 추이",
                       anchor="w", fill=TEXT, font=(FONT, 11, "bold"))

        hist = sorted(getattr(self, "_daily_cache", []),
                      key=lambda h: (h.get("date", ""), h.get("time", "")))
        hist = hist[-24:]
        if not hist:
            cv.create_text(W // 2, H // 2, text="기록 없음 — 원가비교를 실행하면 그래프가 표시됩니다.",
                           fill=MUTED, font=(FONT, 10))
            return

        BAR_PASTEL = "#f6b0b0"   # 파스텔 빨강(막대)
        LABEL_RED = "#c0504d"    # 수치/비율 라벨(가독성 있는 진한 파스텔 빨강)

        mL, mR, mT, mB = sc(40), sc(16), sc(48), sc(34)
        plot_w = max(1, W - mL - mR)
        plot_h = max(1, H - mT - mB)
        base_y = H - mB
        vals = [int(h.get("diff", 0) or 0) for h in hist]
        tots = [int(h.get("total", 0) or 0) for h in hist]
        maxv = max(vals + [1])

        # y축 눈금(0 / 중간 / 최대) + 옅은 그리드라인
        for frac in (0.0, 0.5, 1.0):
            y = base_y - plot_h * frac
            cv.create_line(mL, y, W - mR, y,
                           fill=(BORDER if frac else MUTED), width=1)
            cv.create_text(mL - sc(6), y, text=f"{round(maxv * frac):,}",
                           anchor="e", fill=MUTED, font=(FONT, 8))

        n = len(hist)
        slot = plot_w / n
        bar_w = min(slot * 0.62, sc(34))
        lbl_step = max(1, (n + 11) // 12)   # 날짜 라벨 과밀 방지
        for i, h in enumerate(hist):
            v = vals[i]
            pct = round(v * 100 / tots[i]) if tots[i] else 0
            xc = mL + slot * (i + 0.5)
            bh = plot_h * v / maxv
            top = base_y - bh
            cv.create_rectangle(xc - bar_w / 2, top, xc + bar_w / 2, base_y,
                                fill=BAR_PASTEL, outline="")
            if bar_w >= sc(12) or v == maxv:
                # 막대 위에 수치(굵게) + 비율(%)을 두 줄로 표시
                cv.create_text(xc, top - sc(16), text=f"{v:,}",
                               fill=LABEL_RED, font=(FONT, 8, "bold"))
                cv.create_text(xc, top - sc(6), text=f"{pct}%",
                               fill=LABEL_RED, font=(FONT, 7))
            if i % lbl_step == 0 or i == n - 1:
                d = str(h.get("date", ""))[5:]   # MM-DD
                cv.create_text(xc, base_y + sc(11), text=d,
                               fill=MUTED, font=(FONT, 8))

    def _export_daily(self) -> None:
        hist = list(getattr(self, "_daily_cache", []))
        if not hist:
            pmsg.showwarning("데이터 없음", "내보낼 일일현황 기록이 없습니다.")
            return
        hist.sort(key=lambda h: (h.get("date", ""), h.get("time", "")), reverse=True)
        rows = [{
            "일자": h.get("date", ""), "갱신시각": h.get("time", ""),
            "전체": h.get("total", 0), "원가차이": h.get("diff", 0),
            "미입고": h.get("nostock", 0), "미매칭": h.get("unmatched", 0),
            "원가일치": h.get("same", 0),
        } for h in hist]
        export_rows_excel(rows, "일일현황.xlsx")

    def _prices_updated(self, rows: list[dict], new_all: list[dict], n_fetched: int) -> None:
        """백그라운드 입고단가 매칭 완료 → 전체 데이터 교체 후 현재 필터/정렬로 재표시."""
        if rows is not self._inventory_rows:   # 그 사이 새 조회가 있었으면 무시
            return
        self._inventory_display_all = new_all
        self._render_inventory()


def run_install_window(url: str, ver: str, target: str) -> None:
    """별도 업데이터 창(실행 팝업과 동일한 둥근 프리미엄 디자인).

    하나의 둥근 창에서 다운로드(0~60%) → 설치(60~100%)를 하단 진행률 바로 보여준다.
    Inno 설치는 /VERYSILENT(설치 창 없음)로 실행하고, 완료되면 새 버전을 실행한다.
    이 프로세스는 교체 대상 exe 가 아닌 임시 복사본이라 설치 중에도 살아 있다.
    """
    import subprocess
    root = tk.Tk()
    init_ui_scale(root)
    root.withdraw()
    sp = Splash(root, status="업데이트 준비 중…")
    st = {"installer": None, "installed": False, "failed": False}

    def dl_prog(received, total):
        if total and total > 0:
            pct = received * 100.0 / total
            root.after(0, lambda: (
                sp.set_progress(pct * 0.6),   # 다운로드 = 0~60%
                sp.set_status(f"새 버전 {ver} 다운로드 중… {int(pct)}% "
                              f"({received / 1048576:.1f} / {total / 1048576:.1f} MB)")))

    def worker():
        try:
            path = updater.download_installer(url, progress=dl_prog)
        except Exception:  # noqa: BLE001
            st["failed"] = True
            return
        st["installer"] = path
        root.after(0, lambda: sp.set_status("설치 중… 파일을 적용하고 있습니다"))
        try:
            p = subprocess.Popen([path, "/VERYSILENT", "/SUPPRESSMSGBOXES", "/NORESTART"])
            p.wait()
        except Exception:  # noqa: BLE001
            st["failed"] = True
            return
        st["installed"] = True

    threading.Thread(target=worker, daemon=True).start()

    prog = {"v": 0.0}

    def tick():
        if st["failed"]:
            sp.set_status("업데이트 실패 — 다음 실행 시 다시 시도합니다")
            root.after(1600, root.destroy)
            return
        if st["installed"]:
            sp.set_progress(100)
            sp.set_status("설치 완료 — 새 버전을 실행합니다…")
            if target and os.path.exists(target):
                try:
                    os.startfile(target)  # noqa: S606
                except Exception:  # noqa: BLE001
                    pass
            root.after(900, root.destroy)
            return
        if st["installer"]:        # 설치 단계: 60 → 97 부드럽게
            prog["v"] = min(97.0, max(prog["v"], 60.0) + 1.4)
            sp.set_progress(prog["v"])
        root.after(170, tick)

    root.after(300, tick)
    root.mainloop()


def spawn_install_updater(url: str, ver: str, target: str) -> bool:
    """교체 대상이 아닌 임시 복사본으로 별도 업데이터 창을 띄운다.

    frozen(.exe): 자기 자신을 %TEMP%\\PriceCheckUpdater.exe 로 복사해 실행
                  (설치 시 taskkill 대상인 EcountInventory.exe 와 이름이 달라 살아남음).
    dev(.py)    : python 으로 gui.py --install 실행.
    성공 시 True.
    """
    import subprocess
    import tempfile
    import shutil
    DETACHED = 0x00000008 | 0x00000200   # DETACHED_PROCESS | CREATE_NEW_PROCESS_GROUP
    try:
        if getattr(sys, "frozen", False):
            tmp = os.path.join(tempfile.gettempdir(), "PriceCheckUpdater.exe")
            shutil.copy2(sys.executable, tmp)
            cmd = [tmp, "--install", url, ver, target]
        else:
            cmd = [sys.executable, os.path.abspath(__file__), "--install", url, ver, target]
        subprocess.Popen(cmd, creationflags=DETACHED, close_fds=True)
        return True
    except Exception:  # noqa: BLE001
        return False


def main() -> None:
    enable_dpi_awareness()   # 고DPI에서 흐릿함 방지(Tk 생성 전)
    if len(sys.argv) >= 2 and sys.argv[1] == "--install":
        url = sys.argv[2] if len(sys.argv) > 2 else ""
        ver = sys.argv[3] if len(sys.argv) > 3 else ""
        target = sys.argv[4] if len(sys.argv) > 4 else ""
        run_install_window(url, ver, target)
        return
    App().mainloop()


if __name__ == "__main__":
    main()
